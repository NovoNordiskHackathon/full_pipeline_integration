import pandas as pd
import argparse
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration loader for rules
def load_config(config_path: str) -> dict:
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

"""
Recreate the original template structure with Unnamed columns so that the
resulting DataFrame (before Excel export) matches Final_study_specific_form_initial.py exactly.
"""
template_data = {
    'CTDM to fill in': [None, 'Source', 'New or Copied from Study', None],
    'CTDM Optional, if blank CDP to propose': [None, 'Form', 'Form Label', None],
    'Input needed from SDTM': [None, None, 'Form Name (provided by SDTM Programmer, if SDTM linked form)', None],
    'CDAI input needed': [None, 'Item Group', 'Item Group (if only one on form, recommend same as Form Label)', None],
    'Unnamed: 4': [None, None, 'Item group Repeating', None],
    'Unnamed: 5': [None, None, 'Repeat Maximum, if known, else default =50', None],
    'Unnamed: 6': [None, None, 'Display format of repeating item group (Grid, read only, form)', None],
    'Unnamed: 7': [None, None, 'Default Data in repeating item group', None],
    'Unnamed: 8': [None, 'Item', 'Item Order', None],
    'Unnamed: 9': [None, None, 'Item Label', None],
    'Unnamed: 10': [None, None, 'Item Name (provided by SDTM Programmer, if SDTM linked)', None],
    'Unnamed: 11': [None, 'Progressive Display', None, None],
    'Unnamed: 12': [None, None, 'Progressively Displayed?', None],
    'Unnamed: 13': [None, None, 'Controlling Item (item name if known, else item label)', None],
    'Unnamed: 14': [None, None, 'Controlling Item Value', None],
    'Unnamed: 15': [None, None, None, None],
    'Unnamed: 16': [None, 'Data type', None, None],
    'Unnamed: 17': ['Data type', 'If text or number, Field Length', None, None],
    'Unnamed: 18': [None, 'If number, Precision (decimal places)', None, None],
    'Unnamed: 19': [None, 'Codelist', 'Codelist - Choice Labels (If many, can use Codelist Name)', None],
    'Unnamed: 20': [None, None, 'Codelist Name (provided by SDTM programmer)', None],
    'Unnamed: 21': [None, None, 'Choice Code (provided by SDTM programmer)', None],
    'Unnamed: 22': [None, None, 'Codelist: Control Type', None],
    'Unnamed: 23': [None, 'System Queries', 'If Number, Range: Min Value - Max Value', None],
    'Unnamed: 24': [None, None, 'If Date, Query Future Date', None],
    'Unnamed: 25': [None, None, 'Required', None],
    'Unnamed: 26': [None, None, 'If Required, Open Query when Intentionally Left Blank(Form, Item)', None],
    'Unnamed: 27': [None, None, 'Notes', None]
}

# In-memory template matching the original script
df_template = pd.DataFrame(template_data)




import csv
import re
import pandas as pd


def get_text(node):
    """
    Extract text from a node safely and recursively.
    This can find text nested inside other nodes (e.g., P -> StyleSpan).
    """
    if not isinstance(node, dict):
        return ""
    text = (node.get("text") or "").strip()
    if text:
        return text
    for child in node.get("children", []):
        text = get_text(child)
        if text:
            return text
    return ""




def is_valid_form_name(text):
    """Check if text is a valid form name using strict regex patterns."""
    if not text:
        return False

    form_name_pattern = re.compile(
        r'(?:'
        r'\[([A-Z0-9_\-]{3,})\]'  # Brackets with ALL CAPS, at least 3 chars
        r'|'
        r'.*\b(Non-)?[Rr]epeating\b.*'  # Contains repeating/non-repeating
        r')',
        re.IGNORECASE
    )

    match = form_name_pattern.search(text)
    if not match:
        return False

    if match.group(1):  # Bracketed match
        bracketed_content = match.group(1)

        # ADDED FROM VERSION 1: Enforce uppercase requirement
        if not bracketed_content.isupper():
            return False

        invalid_bracketed_patterns = [
            r'^L\d+$',  # L1, L2, etc.
            r'^[A-Z]\d+$',  # A1, B2, etc.
            r'^A\d+$',  # A200, etc.
        ]
        for pattern in invalid_bracketed_patterns:
            if re.match(pattern, bracketed_content):
                return False
        return True

    # FROM VERSION 2: Explicit elif for repeating pattern
    elif 'repeating' in text.lower():
        if len(text) < 10 or len(text) > 80:
            return False
        exclusion_patterns = [
            r'^(CRF|Form)\s+(Date|Time|Coordinator|Designer|Notes?).*',
            r'^\w{1,4}\s+(Date|Time|Coordinator|Designer)\b.*',
            r'^\s*(Date|Time|Coordinator|Designer)\s*-\s*(Non-)?[Rr]epeating.*',
        ]
        for pattern in exclusion_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return False
        return True

    # FROM VERSION 2: Explicit return False
    return False


def is_valid_form_label(text):
    """Check if text is a valid form label."""
    # FROM VERSION 1: Check both min and max length upfront
    if not text or len(text) < 3 or len(text) > 100:
        return False

    invalid_patterns = [
        r'^\s*V\d+[A-Z]*\s*$',  # Just visit numbers
        r'Design\s*Notes?\s*:?$',
        r'Oracle\s*item\s*design\s*notes?\s*:?$',
        r'General\s*item\s*design\s*notes?\s*:?$',
        r'^\s*Non-Visit\s*Related\s*$',
        r'^Data from.*',
        r'^Hidden item.*',
        r'^The item.*',
        r'^\d+\s+',
        r'.*\|A\d+\|.*',
        r'^\s*(Non-)?[Rr]epeating(\s+form)?\s*$',
    ]
    for pattern in invalid_patterns:
        if re.match(pattern, text, re.IGNORECASE):
            return False
    return True


def extract_forms_cleaned(data):
    """Extract forms with improved duplicate handling and validation."""
    results = []
    seen_forms = set()

    def process_h1_section(h1_node):
        h1_text = get_text(h1_node)
        if not is_valid_form_label(h1_text):
            h1_text = "Unknown Section"

        def find_forms_in_node(node, current_label=None):
            if not isinstance(node, dict):
                return
            node_name, node_text = node.get("name", ""), get_text(node)

            if node_name.startswith("H2") and is_valid_form_label(node_text) and not is_valid_form_name(node_text):
                current_label = node_text

            if is_valid_form_name(node_text):
                form_name, form_label = node_text, current_label if current_label else h1_text
                form_key = (form_label, form_name)
                if form_key not in seen_forms:
                    results.append({
                        "Form Label": form_label,
                        "Form Name": form_name,
                        "H1_Text": h1_text,
                        "Form_Node": node,
                        "Parent_H1_Node": h1_node
                    })
                    seen_forms.add(form_key)

            for child in node.get("children", []):
                find_forms_in_node(child, current_label)

        find_forms_in_node(h1_node, None)

    def find_h1_sections(node):
        if not isinstance(node, dict):
            return
        if node.get("name", "").startswith("H1"):
            process_h1_section(node)
        for child in node.get("children", []):
            find_h1_sections(child)

    find_h1_sections(data)
    return results


def find_nodes_by_name_pattern(node, pattern):
    """Find all nodes matching a name pattern recursively."""
    if not isinstance(node, dict):
        return []
    matches = []
    if re.search(pattern, node.get("name", "")):
        matches.append(node)
    for child in node.get("children", []):
        matches.extend(find_nodes_by_name_pattern(child, pattern))
    return matches


# ==============================================================================
# 🔥 CORRECTED ITEM EXTRACTION LOGIC
# ==============================================================================

# ============== NEW HELPER FUNCTIONS - ADD THESE ==================

def is_metadata_table(table_node):
    """
    🔥 FIXED: Detect and skip metadata/header tables containing document information.
    Uses internal recursive text collection to avoid modifying get_text() used elsewhere.
    These tables typically contain: company name, Trial ID, Date, Version, Page numbers, etc.
    """
    if not isinstance(table_node, dict):
        return False

    # 🔥 NEW: Internal function to collect ALL text from table (not affecting get_text())
    def get_all_table_text(node):
        """
        Internal helper to recursively collect ALL text from a node and its children.
        This is used ONLY for metadata detection and doesn't affect other code.
        """
        if not isinstance(node, dict):
            return ""

        text_parts = []

        # Get text from current node
        if node.get("text"):
            text_parts.append(node.get("text").strip())

        # Recursively get text from ALL children
        for child in node.get("children", []):
            child_text = get_all_table_text(child)
            if child_text:
                text_parts.append(child_text)

        # Join all text parts with space
        return " ".join(text_parts)

    # Get ALL text from the table using internal function
    table_text = get_all_table_text(table_node)

    # Define metadata keywords (configurable)
    metadata_keywords = CONFIG.get('metadata_keywords', [
        r'Novo\s+Nordisk',
        r'Trial\s+ID\s*:',
        r'Sample\s+eCRF',
        r'Mock-up',
        r'requirement',
        r'Version\s*:\s*\d+\.\d+',
        r'Page\s*:\s*\d+\s+of\s+\d+',
    ])

    # Count how many metadata patterns are found
    matches = 0
    for pattern in metadata_keywords:
        if re.search(pattern, table_text, re.IGNORECASE):
            matches += 1

    # If 3 or more metadata patterns found, it's likely a metadata table
    if matches >= 3:
        return True

    # Additional check: Look for specific company/organization names
    company_patterns = CONFIG.get('company_patterns', [
        r'Novo\s+Nordisk\s+A/S',
        r'Clinical\s+Trial',
        r'Protocol',
    ])

    for pattern in company_patterns:
        if re.search(pattern, table_text, re.IGNORECASE):
            # If company name found + at least one other metadata field, skip it
            if matches >= 2:
                return True

    return False


def check_p_extracharspan_extracharspan_pattern(node):
    """
    Check if node has pattern: P -> ExtraCharSpan -> ExtraCharSpan[]
    Returns True if a P node contains ExtraCharSpan, which itself contains ExtraCharSpan(s)
    """
    if not isinstance(node, dict):
        return False

    # Check if current node is P
    if node.get("name", "") == "P":
        children = node.get("children", [])
        for child in children:
            # Check if child is ExtraCharSpan
            if child.get("name", "") == "ExtraCharSpan":
                # Check if this ExtraCharSpan has ExtraCharSpan children
                grandchildren = child.get("children", [])
                for grandchild in grandchildren:
                    if grandchild.get("name", "") == "ExtraCharSpan":
                        return True

    # Recursively check all children
    for child in node.get("children", []):
        if check_p_extracharspan_extracharspan_pattern(child):
            return True

    return False


def check_p_sub_pattern(node):
    """
    Check if node has pattern: P -> Sub
    Returns True if a P node directly contains a Sub child
    """
    if not isinstance(node, dict):
        return False

    # Check if current node is P
    if node.get("name", "") == "P":
        children = node.get("children", [])
        for child in children:
            # Check if child is Sub
            if child.get("name", "") == "Sub":
                return True

    # Recursively check all children
    for child in node.get("children", []):
        if check_p_sub_pattern(child):
            return True

    return False


# ================================================================

def is_instruction(text):
    """
    Check if text is likely an instruction based on keywords and punctuation density,
    including 'collect' and 'integration'.
    """
    if not isinstance(text, str) or not text.strip():
        return False
    text = text.strip()

    # 🔥 SOLUTION: A sentence ending in '?' is a question, not an instruction.
    # This check will now correctly identify "Have blood samples been collected?" as a question.
    if text.endswith('?'):
        return False

    # Rule 1: Keywords ('collect' and 'integration' are included)
    kw_list = CONFIG.get('instruction_keywords', [
        'please','note','ensure','click','enter','complete','select','indicate','check','provide','collect','integration','Study ID'
    ])
    keywords = re.compile(r'\b(' + '|'.join(map(re.escape, kw_list)) + r')\b', re.IGNORECASE)

    # If the text contains 'integration' or any other keyword, it is an instruction.
    if keywords.search(text):
        return True

    # Rule 2: Punctuation density (a rough heuristic)
    punctuation_count = len(re.findall(r'[:\-\(\)\.?!;]', text))
    word_count = len(text.split())

    if word_count < 5 and punctuation_count >= 1:
        return True
    elif word_count >= 5 and word_count > 0 and (punctuation_count / word_count) > 0.1:
        return True

    # Rule 3: Start with a number and period (list/step instruction)
    if re.match(r'^\s*\d+\.\s+\w', text):
        return True

    return False

def is_valid_option_content(node):
    """
    🔥 ENHANCED: Check if a TD node contains valid option content (not metadata/annotations).

    Filters out false positives like:
    - All capital letters: 'CO', 'RT', 'R', 'C'
    - Comma-separated capital letters: 'C, CO', 'A, R', 'A,R', 'A, R, CO, RT'

    Returns True for valid options like:
    - Mixed case words: 'Yes', 'No', 'Maybe'
    - Longer text: 'Date format', 'Select one'
    - Numbers: '1', '2', '3'
    """
    if not isinstance(node, dict):
        return False

    # Get text from the node
    node_text = get_text(node).strip()

    if not node_text:
        return False

    # 🔥 NEW Rule 1: Check if text contains ONLY capital letters, commas, and spaces
    # This catches: "CO", "RT", "C", "R", "C, CO", "A, R, CO, RT"
    only_caps_comma_space = bool(re.match(r'^[A-Z,\s]+$', node_text))

    if only_caps_comma_space:
        # If it's all caps, reject it (likely metadata/annotation)
        return False

    # 🔥 Rule 2: Additional pattern check - single/double letter codes separated by commas
    # Pattern: X, XX or X,XX (e.g., "C, CO", "A,R")
    # This is a redundant check but kept for extra safety
    short_code_pattern = r'^[A-Z]{1,2}(\s*,\s*[A-Z]{1,2})+$'
    if re.match(short_code_pattern, node_text):
        return False

    # If it passes all filters, it's likely valid option content
    # This will allow: "Yes", "No", "Maybe", "Date", "1", "2", etc.
    return True








def has_option_child(node):
    """
    Recursively check if a node contains an option-indicating child.
    Checks three patterns in order:
    1. Direct option nodes (LI, L, ExtraCharSpan, LBody)
    2. P/ExtraCharSpan/ExtraCharSpan[] pattern
    3. P/Sub pattern
    """
    if not isinstance(node, dict):
        return False

    # LOGIC 1: Original check for direct option-indicating nodes
    if node.get("name", "") in ["LI", "L", "ExtraCharSpan", "LBody"]:
        return True

    for child in node.get("children", []):
        if has_option_child(child):
            return True

    # LOGIC 2: Check for P/ExtraCharSpan/ExtraCharSpan[] pattern
    if check_p_extracharspan_extracharspan_pattern(node):
        return True

    # LOGIC 3: Check for P/Sub pattern
    if check_p_sub_pattern(node):
        return True

    # 🔥 NEW LOGIC 4: If TD contains P nodes with text, treat as option cell
    # This catches cases like "|A3| RT" and "|N3| RT" that were being missed
    if node.get("name", "").startswith("TD"):
        # 🔥 NEW: First check if this TD contains valid option content
        if not is_valid_option_content(node):
            return False
        p_nodes = find_nodes_by_name_pattern(node, r'^P')
        if p_nodes:
                # Check if any P node has meaningful text (not just whitespace)
                for p_node in p_nodes:
                    p_text = get_text(p_node).strip()
                    if p_text and len(p_text) > 0:
                        return True

    return False


def extract_items_from_form(form_node):
    """
    Extracts item data, handling rows with TH (question) + TD (options),
    and persistently tracking the Item Group across table breaks.
    """
    items_data = []
    table_nodes = find_nodes_by_name_pattern(form_node, r'^Table')

    # 🔥 FIX: Initialize Item Group outside the table loop to persist across table breaks.
    current_item_group = ""

    for table in table_nodes:
        # 🔥 NEW: Skip metadata tables
        if is_metadata_table(table):
            print(f"⚠️  Skipping metadata table: {table.get('name', '')}")
            continue
        tr_nodes = find_nodes_by_name_pattern(table, r'^TR')

        for tr in tr_nodes:
            # 🔥 Get ALL TH and TD cells in a row
            cells = [child for child in tr.get("children", []) if child.get("name", "").startswith(("TH", "TD"))]

            # 🔥 ITEM GROUP LOGIC: Check for a single-cell row that is likely an Item Group header
            if len(cells) == 1:
                potential_group_text = get_text(cells[0])
                # Only treat it as an Item Group if it is a valid label and NOT an instruction
                if is_valid_form_label(potential_group_text) and not is_instruction(potential_group_text):
                    current_item_group = potential_group_text
                    continue  # Skip processing this row as an item
            # 🔥 END ITEM GROUP LOGIC

            # 🔥 NEW: Handle 3-column structure (TH | TD | TD[2])
            if len(cells) == 3:
                th_cell, question_cell, option_cell = cells

                # Check if this is a valid question row (TH has asterisk, question_cell has text)
                th_text = get_text(th_cell).strip()
                question_text = ""

                # Extract question from TD (second column)
                p_nodes = find_nodes_by_name_pattern(question_cell, r'^P')
                if p_nodes:
                    # Check if ALL P nodes are ParagraphSpan (skip category headers)
                    all_paragraph_spans = all(node.get("name", "").startswith("ParagraphSpan") for node in p_nodes)
                    if all_paragraph_spans:
                        continue

                    # Extract text from ALL P nodes
                    p_texts = [get_text(p_node) for p_node in p_nodes if get_text(p_node)]
                    question_text = "\n".join(p_texts)
                else:
                    question_text = get_text(question_cell)

                    # 🔥 ADD THIS LINE HERE - RIGHT AFTER EXTRACTING question_text
                if not question_text or not question_text.strip() or question_text.strip() in ["*", "**", "***"]:
                    continue

                # 🔥 CRITICAL FIX 1: Check if the question text is an instruction
                if is_instruction(question_text):
                    print(f"    ⚠️  Skipping instruction row (3-col): '{question_text}'")
                    continue


                # 🔥 NEW: Check if option_cell contains valid option content
                # Skip rows where the option cell has metadata like "C, CO"
                if not is_valid_option_content(option_cell):
                    print(f"    ⚠️  Skipping false positive: '{get_text(option_cell)}' (metadata/annotation)")
                    continue

                # # 🔥 ENHANCED: Combine TH text with question text if TH contains "*" or meaningful prefix
                # # This handles cases where "*" is in a separate TH column
                # if th_text and th_text in ["*", "**", "***"]:
                #     # Prepend the asterisk to the question text
                #     question_text = f"{th_text} {question_text}"
                # elif th_text and len(th_text) <= 10:  # Short prefix (like a number or code)
                #     # Prepend the prefix
                #     question_text = f"{th_text} {question_text}"



                # Add this item (third column becomes the option node)
                items_data.append({
                    "Item Group": current_item_group,  # 🔥 ASSIGN ITEM GROUP
                    "Item Name": question_text,
                    "Option_TD_Node": option_cell
                })
                continue

            # 🔥 ORIGINAL LOGIC: Handle 2-column structure (TH/TD | TD with options)
            for i, cell in enumerate(cells):
                if i > 0 and has_option_child(cell):
                    prev_cell = cells[i - 1]
                    item_name_text = ""
                    # 🔥 NEW: Extract from Sub nodes first (for TH cells with Sub children)
                    sub_nodes = find_nodes_by_name_pattern(prev_cell, r'^Sub')
                    if sub_nodes:
                        # Get the first Sub node (the actual label, not [hidden]/[read-only])
                        main_sub_text = get_text(sub_nodes[0]).strip()
                        if main_sub_text and not main_sub_text.startswith('['):
                            item_name_text = main_sub_text

                    # If no Sub nodes or Sub extraction failed, try P nodes
                    if not item_name_text:

                        p_nodes = find_nodes_by_name_pattern(prev_cell, r'^P')

                        if p_nodes:
                            all_paragraph_spans = all(node.get("name", "").startswith("ParagraphSpan") for node in p_nodes)
                            if all_paragraph_spans:
                                continue

                            p_texts = [get_text(p_node) for p_node in p_nodes if get_text(p_node)]
                            item_name_text = "\n".join(p_texts)
                        else:
                            item_name_text = get_text(prev_cell)

                        # 🔥 NEW: Skip if item_name_text is ONLY asterisks
                        if not item_name_text or item_name_text.strip() in ["*", "**", "***"]:
                            continue
                     # 🔥 CRITICAL FIX 2: ADD THIS INSTRUCTION CHECK!
                    if is_instruction(item_name_text):
                        print(f"    ⚠️  Skipping instruction row (2-col): '{item_name_text}'")
                        continue


                    items_data.append({
                        "Item Group": current_item_group,  # 🔥 ASSIGN ITEM GROUP
                        "Item Name": item_name_text,
                        "Option_TD_Node": cell
                    })

    # 🔥 Enhanced deduplication using Item Group + Item Name
    unique_items = []
    seen_names = set()
    for item in items_data:
        item_key = (item["Item Name"], item.get("Item Group", ""))  # 🔥 Tuple key for uniqueness
        if item_key not in seen_names:
            seen_names.add(item_key)
            unique_items.append(item)
    return unique_items


def determine_data_type(option_td_node, codelist_content):
    """
    Determine data type based on:
    1. Codelist content patterns (Date/Time, Label)
    2. JSON node structure (Codelist)
    3. Default to Text

    Parameters:
    - option_td_node: The TD node containing options from JSON
    - codelist_content: The text content from "Codelist - Choice Labels" column
    """
    if not option_td_node:
        return "Text"

    # Ensure codelist_content is a string
    if codelist_content is None:
        codelist_content = ""
    else:
        codelist_content = str(codelist_content).strip()

    # 🔥 LOGIC 1: Check for Date/Time pattern in codelist content
    # Pattern: Req/Req/Req(YYYY-YYYY) or similar date range patterns
    date_time_pattern = CONFIG.get('date_time_pattern', r'Req.*?\(\d{4}[-–—/]{1,2}\d{4}\)')
    if re.search(date_time_pattern, codelist_content, re.IGNORECASE):
        return "Date/Time"

    # 🔥 LOGIC 2: Check for Codelist in JSON structure
    # Look for LBody nodes that contain ExtraCharSpan children
    lbody_nodes = find_nodes_by_name_pattern(option_td_node, r'^LBody')
    if lbody_nodes:
        for lbody_node in lbody_nodes:
            # Check if this LBody has ExtraCharSpan children
            extracharspan_nodes = find_nodes_by_name_pattern(lbody_node, r'^ExtraCharSpan')
            if extracharspan_nodes:
                return "Codelist"

    # Also check for direct ExtraCharSpan nodes (original logic)
    if find_nodes_by_name_pattern(option_td_node, r'^ExtraCharSpan'):
        return "Codelist"

    # 🔥 LOGIC 3: Check for Label pattern
    # Pattern: Contains |...| but NO multiple bullet points (no multiple •)
    # This catches: • |N3| Years, • |0 < N3 ≤ 200| ¡ kg
    if '|' in codelist_content:
        # Count bullet points
        bullet_count = codelist_content.count('•')
        # If only ONE bullet (or none) and contains pipes, it's a Label
        if bullet_count <= 1:
            return "Label"

    # 🔥 LOGIC 4: Default to Text
    return "Text"


def get_all_lbody_values(option_td_node):
    """
    Get all option values from the specific option cell.
    Extracts from LBody, Sub, or P nodes depending on the structure.
    """
    if not option_td_node:
        return ""

    # First try to find LBody nodes (for radio button/codelist options)
    lbody_nodes = find_nodes_by_name_pattern(option_td_node, r'^LBody')
    if lbody_nodes:
        values = [get_text(node) for node in lbody_nodes if get_text(node)]
        seen = set()
        unique_values = [x for x in values if not (x in seen or seen.add(x))]
        return "\n".join(f"• {val}" for val in unique_values)

    # 🔥 NEW: Try to find Sub nodes (for subscript-style options)
    sub_nodes = find_nodes_by_name_pattern(option_td_node, r'^Sub')
    if sub_nodes:
        values = []
        for node in sub_nodes:
            text = get_text(node)
            # Skip empty text and special characters
            if text and text.strip() not in ["", "\uf0fe", "□", "¡"]:
                # Clean up the text (remove leading bullets/symbols)
                cleaned_text = text.strip().lstrip("¡ ").lstrip("□ ").strip()
                if cleaned_text:
                    values.append(cleaned_text)

        seen = set()
        unique_values = [x for x in values if not (x in seen or seen.add(x))]
        if unique_values:
            return "\n".join(f"• {val}" for val in unique_values)

    # Last resort: Try P nodes (for date/text format fields)
    p_nodes = find_nodes_by_name_pattern(option_td_node, r'^P')
    if p_nodes:
        values = []
        for node in p_nodes:
            text = get_text(node)
            # Skip empty text and special characters
            if text and text.strip() not in ["", "\uf0fe", "□", "¡"]:
                values.append(text)

        seen = set()
        unique_values = [x for x in values if not (x in seen or seen.add(x))]
        if unique_values:
            return "\n".join(f"• {val}" for val in unique_values)

    return ""


def calculate_field_length(codelist_content):
    """
    Calculate field length from codelist content for Text or Label data types.
    Extracts the maximum length from patterns like |N3| (3 characters) or |0 < N3 ≤ 200| (3 digits).
    """
    if not codelist_content:
        return ""

    # Clean up the content
    content = str(codelist_content).strip()

    # Pattern 1: |Nxx| where xx is the number of characters/digits (e.g., |N3| = 3)
    simple_n_pattern = r'\|N(\d+)\|'
    match = re.search(simple_n_pattern, content)
    if match:
        return match.group(1)  # Returns the digit (e.g., "3" from |N3|)

    # Pattern 2: |0 < N3 ≤ 200| - extract the digit from N3
    complex_n_pattern = r'\|.*N(\d+).*\|'
    match = re.search(complex_n_pattern, content)
    if match:
        return match.group(1)  # Returns the digit (e.g., "3" from N3)

    # Pattern 3: For plain text, calculate the actual length of content (excluding bullets)
    # Remove bullet points and calculate max line length
    lines = content.split('\n')
    max_length = 0
    for line in lines:
        cleaned_line = line.strip().lstrip('• ').strip()
        if cleaned_line and not cleaned_line.startswith('|'):  # Ignore format patterns
            max_length = max(max_length, len(cleaned_line))

    if max_length > 0:
        return str(max_length)

    return ""


def calculate_precision(codelist_content):
    """
    Calculate precision (decimal places) from codelist content for Label data types.
    Extracts precision from patterns like |N3.2| (2 decimal places).
    """
    if not codelist_content:
        return ""

    content = str(codelist_content).strip()

    # Pattern 1: |N3.2| where 3 is total digits and 2 is decimal places
    precision_pattern = r'\|N\d+\.(\d+)\|'
    match = re.search(precision_pattern, content)
    if match:
        return match.group(1)  # Returns the decimal places (e.g., "2" from |N3.2|)

    # Pattern 2: Look for decimal notation in complex patterns like |0.00 < N3.2 ≤ 200.00|
    decimal_in_constraint = r'\|.*N\d+\.(\d+).*\|'
    match = re.search(decimal_in_constraint, content)
    if match:
        return match.group(1)

    # Pattern 3: Check if there are decimal values in the range (e.g., 0.00, 200.00)
    decimal_values = re.findall(r'\d+\.(\d+)', content)
    if decimal_values:
        # Return the maximum decimal places found
        max_decimals = max(len(d) for d in decimal_values)
        return str(max_decimals)

    # Default: No decimal places (integer)
    return "0"


def extract_number_range(codelist_content):
    """
    Extract numeric range from patterns like |0 < N3 ≤ 200| → "0 - 200"
    """
    if not codelist_content:
        return ""

    content = str(codelist_content).strip()

    # Pattern: |min < Nx ≤ max| or |min ≤ Nx ≤ max| or |min < Nx < max|
    range_pattern = r'\|(\d+(?:\.\d+)?)\s*[<≤]\s*N\d+(?:\.\d+)?\s*[<≤]\s*(\d+(?:\.\d+)?)\|'
    match = re.search(range_pattern, content)
    if match:
        min_val = match.group(1)
        max_val = match.group(2)
        return f"{min_val} - {max_val}"

    # Pattern: |0 < N3| (only minimum, no maximum)
    min_only_pattern = r'\|(\d+(?:\.\d+)?)\s*[<≤]\s*N\d+(?:\.\d+)?\|'
    match = re.search(min_only_pattern, content)
    if match:
        min_val = match.group(1)
        return f"{min_val} - "

    # Pattern: |N3 ≤ 200| (only maximum, no minimum)
    max_only_pattern = r'\|N\d+(?:\.\d+)?\s*[<≤]\s*(\d+(?:\.\d+)?)\|'
    match = re.search(max_only_pattern, content)
    if match:
        max_val = match.group(1)
        return f" - {max_val}"

    return ""


def check_query_future_date(data_type):
    """
    Determine if future dates should trigger a query.
    For Date/Time fields, typically set to "No" (allow future dates).
    """
    if data_type == "Date/Time":
        return "Y"
    return ""


def check_required_field(item_name, sibling_cells=None, current_cell_index=None):
    """
    Check if item is required by:
    - Looking for asterisk (*) at the start of item_name.
    - Also checking for '*' in the previous cell's text (direct sibling cell before the current cell).

    Parameters:
    - item_name: text of current item label
    - sibling_cells: list of TD/TH nodes (cells) in the row
    - current_cell_index: index of the cell with item_name in sibling_cells
    """
    if not item_name and sibling_cells is None:
        return "N"

    item_text = str(item_name).strip() if item_name else ""

    if item_text.startswith('*'):
        return "Y"
    return "N"




# ==============================================================================
# NEW HELPER FUNCTIONS FOR ITEM GROUP REPEATING LOGIC
# ==============================================================================

def analyze_item_groups_per_form(items):
    """
    Analyze item groups for a form to determine which ones are repeating.

    Parameters:
    - items: List of item dictionaries with 'Item Group' and 'Item Name' keys

    Returns:
    - item_group_counts: Dict mapping each Item Group to its occurrence count
    - repeating_groups: Set of Item Groups that occur more than once
    """
    from collections import Counter

    # Extract all item groups (excluding empty ones)
    item_groups = [
        item.get("Item Group", "").strip()
        for item in items
        if item.get("Item Group", "").strip() and item.get("Item Group", "").strip() != 'NaN'
    ]

    # Count occurrences of each item group
    item_group_counts = Counter(item_groups)

    # Identify repeating groups (count > 1)
    repeating_groups = {group for group, count in item_group_counts.items() if count > 1}

    return item_group_counts, repeating_groups


def get_item_group_repeating_flag(item_group, repeating_groups):
    """
    Determine if an item group is repeating.

    Parameters:
    - item_group: The item group value for the current item
    - repeating_groups: Set of item groups that are repeating

    Returns:
    - 'Y' if the item group is repeating, 'N' otherwise
    """
    if not item_group or item_group == 'NaN' or item_group.strip() == '':
        return 'N'

    return 'Y' if item_group in repeating_groups else 'N'


def get_repeat_maximum(item_group, item_group_repeating_flag, item_group_counts):
    """
    Determine the repeat maximum value for an item.

    Parameters:
    - item_group: The item group value for the current item
    - item_group_repeating_flag: 'Y' or 'N' indicating if the item group is repeating
    - item_group_counts: Dict mapping each Item Group to its occurrence count

    Returns:
    - The count of occurrences if repeating ('Y'), otherwise 50
    """
    if item_group_repeating_flag == 'Y' and item_group in item_group_counts:
        return item_group_counts[item_group]
    else:
        return 50


# ==============================================================================
# UPDATED HELPER FUNCTION FOR ITEM ORDER
# ==============================================================================

def assign_item_order(items):
    """
    Assign sequential order numbers (1, 2, 3...) to items based on their
    appearance order in the Item Label column for the current form.

    Parameters:
    - items: List of item dictionaries with 'Item Name' (Item Label) key

    Returns:
    - items: List with added 'Item_Order' key for each item

    Logic: Items are numbered sequentially as 1, 2, 3, 4... based on their
           order in the form's Item Label column.
    """
    for idx, item in enumerate(items, start=1):
        item['Item_Order'] = idx

    return items


# ==============================================================================
# UPDATED MAIN PROCESSING FUNCTION WITH SIMPLE ITEM ORDER
# ==============================================================================

def process_clinical_forms(json_file_path, template_csv_path=None, output_csv_path="Study_Specific_Form.xlsx", config_path: str = "./config/config_study_specific_forms.json"):
    """Main function to process JSON and create the item-based Excel with repeating logic and item order."""
    global CONFIG
    CONFIG = load_config(config_path)
    # Build template that mirrors the original script (with Unnamed columns)
    template_df = df_template.copy()
    print("✅ Template CSV loaded successfully")

    with open(json_file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
    print("✅ JSON data loaded successfully")

    extracted_forms = extract_forms_cleaned(data)
    print(f"✅ Found {len(extracted_forms)} forms to process")

    all_item_rows = []
    print("\n🔄 Processing forms with item group repeating logic and sequential item order...")

    for form in extracted_forms:
        items = extract_items_from_form(form['Form_Node'])
        print(f"  > Form '{form['Form Name']}': Found {len(items)} unique items.")

        if not items:
            items.append({"Item Name": "", "Option_TD_Node": None, "Item Group": ""})

        # 🔥 UPDATED: Assign sequential item order (1, 2, 3...) based on Item Label sequence
        items = assign_item_order(items)

        # Analyze item groups for this form to determine repeating status
        item_group_counts, repeating_groups = analyze_item_groups_per_form(items)

        print(f"    📊 Item Group Analysis:")
        print(f"       - Total unique item groups: {len(item_group_counts)}")
        print(f"       - Repeating item groups: {len(repeating_groups)}")
        if repeating_groups:
            print(f"       - Repeating groups: {repeating_groups}")
        print(f"    📋 Item Order assigned: {items[0].get('Item_Order', 'N/A')} to {items[-1].get('Item_Order', 'N/A')}")

        for item in items:
            item_row = {}
            option_node = item.get("Option_TD_Node")
            item_name = item['Item Name']

            # Extract Item Group and set to 'NaN' if empty
            item_group_value = item.get("Item Group", "")
            if item_group_value == "":
                item_group_value = 'NaN'

            # Determine if this item group is repeating
            item_group_repeating_flag = get_item_group_repeating_flag(
                item_group_value,
                repeating_groups
            )

            # Calculate repeat maximum based on repeating status
            repeat_maximum = get_repeat_maximum(
                item_group_value,
                item_group_repeating_flag,
                item_group_counts
            )

            # 🔥 Get sequential item order (1, 2, 3...)
            item_order = item.get('Item_Order', 1)

            # Fill in the columns
            item_row['CTDM Optional, if blank CDP to propose'] = form['Form Label']
            item_row['Input needed from SDTM'] = form['Form Name']
            item_row['CDAI input needed'] = item_group_value

            # Fill legacy Unnamed columns exactly like the original implementation
            item_row['Unnamed: 4'] = item_group_repeating_flag
            item_row['Unnamed: 5'] = repeat_maximum
            item_row['Unnamed: 8'] = item_order
            item_row['Unnamed: 9'] = item_name
            item_row['Unnamed: 10'] = ""

            # Get codelist content first
            codelist_content = get_all_lbody_values(option_node)
            item_row['Unnamed: 19'] = codelist_content

            # Determine data type
            data_type = determine_data_type(option_node, codelist_content)
            item_row['Unnamed: 16'] = data_type
            item_row['Unnamed: 22'] = "Radio Button-Vertical" if data_type == "Codelist" else ""

            # Calculate Field Length for Text or Label types
            if data_type in ["Text", "Label"]:
                field_length = calculate_field_length(codelist_content)
                item_row['Unnamed: 17'] = field_length
            else:
                item_row['Unnamed: 17'] = ""

            # Calculate Precision for Label type only
            if data_type == "Label":
                precision = calculate_precision(codelist_content)
                item_row['Unnamed: 18'] = precision
            else:
                item_row['Unnamed: 18'] = ""

            # Extract number range for Label type
            if data_type == "Label":
                number_range = extract_number_range(codelist_content)
                item_row['Unnamed: 23'] = number_range
            else:
                item_row['Unnamed: 23'] = ""

            # Check if future dates should trigger query
            query_future_date = check_query_future_date(data_type)
            item_row['Unnamed: 24'] = query_future_date

            # Check if field is required (based on * in item name)
            is_required = check_required_field(item_name)
            item_row['Unnamed: 25'] = is_required

            # Set "Form,Item" if required, otherwise blank
            if is_required == "Y":
                item_row['Unnamed: 26'] = "Form,Item"
            else:
                item_row['Unnamed: 26'] = ""

            all_item_rows.append(item_row)

    # =========================
    # Build formatted Excel per CTDM 4-row header spec
    # =========================

    # Define grouped headers and colors (light pastels)
    groups = [
        {
            "name": "Source",
            "subheaders": [
                "New or Copied from Study",
            ],
            "color": "E7E6E6",  # light gray
        },
        {
            "name": "Form",
            "subheaders": [
                "Form Label",
                "Form Name (provided by SDTM Programmer, if SDTM linked form)",
            ],
            "color": "C6EFCE",  # light green
        },
        {
            "name": "Item Group",
            "subheaders": [
                "Item Group (if only one on form, recommend same as Form Label)",
                "Item group Repeating",
                "Repeat Maximum, if known, else default =50",
                "Display format of repeating item group (Grid, read only, form)",
                "Default Data in repeating item group",
            ],
            "color": "B3E5FC",  # light blue
        },
        {
            "name": "Item",
            "subheaders": [
                "Item Order",
                "Item Label",
                "Item Name (provided by SDTM Programmer, if SDTM linked item)",
            ],
            "color": "FFD7A8",  # light orange
        },
        {
            "name": "Progressive Display",
            "subheaders": [
                "Progressively displayed?",
                "Controlling item (item triggering it, if yes, describe item below)",
                "Controlling item value",
            ],
            "color": "B3E5FC",  # light blue
        },
        {
            "name": "Data Type",
            "subheaders": [
                "Data type",
                "If text or number, Field Length",
                "If number, Precision (decimal places)",
            ],
            "color": "FFF9C4",  # pale yellow
        },
        {
            "name": "Codelist",
            "subheaders": [
                "Codelist – Choice Labels (if binary, can use Goodlist Table)",
                "Codelist Name (provided by SDTM Programmer)",
                "Choice Code (provided by SDTM Programmer)",
                "Codelist Control Type",
            ],
            "color": "E2F0D9",  # light green variant
        },
        {
            "name": "System Queries",
            "subheaders": [
                "If number, Range: Min Value / Max Value",
                "Date: Query Future Date",
                "Required",
                "If Required, Open Query when intentionally left blank (form/item)",
            ],
            "color": "F8CBAD",  # light orange variant
        },
        {
            "name": "Notes",
            "subheaders": [
                "Notes",
            ],
            "color": "E6B8AF",  # light brownish
        },
    ]

    # Map each group to top CTDM meta category (Row 1)
    ctdm_meta_by_group = {
        "Source": "CTDM to fill in",
        "Form": "CTDM Optional, if blank CDP to propose",
        "Item Group": "CDAI input needed",
        "Item": "Input needed from SDTM",
        "Progressive Display": "CDAI input needed",
        "Data Type": "CDAI input needed",
        "Codelist": "Input needed from SDTM",
        "System Queries": "CDAI input needed",
        "Notes": "CTDM to fill in",
    }

    # Create workbook/sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Study Specific Forms"

    # Styles
    header_font = Font(bold=True, size=10)
    subheader_font = Font(bold=True, size=9)
    ctdm_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Compute total columns
    total_cols = sum(len(g["subheaders"]) for g in groups)

    # Row 1: CTDM meta labels ONCE in A1:D1; E+ blank
    ctdm_titles = [
        "CTDM to fill in",
        "CTDM Optional, if blank CDP to propose",
        "Input needed from SDTM",
        "CDAI input needed",
    ]
    for idx, title in enumerate(ctdm_titles, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = header_font
        cell.alignment = center
        cell.fill = ctdm_fill
    # Ensure E+ remain blank (no styling assigned)

    # Row 2: Group names with merged cells spanning subheaders
    col_start = 1
    for group in groups:
        width = len(group["subheaders"])
        start_col = col_start
        end_col = col_start + width - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        group_fill = PatternFill(start_color=group["color"], end_color=group["color"], fill_type="solid")
        # Set value and style on the merged top-left cell
        c = ws.cell(row=2, column=start_col, value=group["name"])
        c.font = header_font
        c.alignment = center
        c.fill = group_fill
        # Apply fill to the whole merged span to ensure consistent background
        for i in range(start_col, end_col + 1):
            ws.cell(row=2, column=i).fill = group_fill
        col_start += width

    # Row 3: Subheaders (individual cells)
    col_start = 1
    for group in groups:
        width = len(group["subheaders"])
        group_fill = PatternFill(start_color=group["color"], end_color=group["color"], fill_type="solid")
        for i, sub in enumerate(group["subheaders"]):
            c = ws.cell(row=3, column=col_start + i, value=sub)
            c.font = subheader_font
            c.alignment = center
            c.fill = group_fill
        col_start += width

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 28

    # Data rows start at row 4
    start_data_row = 4

    # Helper: safe getter
    def gv(d, key, default=""):
        return d.get(key, default)

    # Assemble rows in the exact subheader order
    ordered_subheaders = []
    for g in groups:
        ordered_subheaders.extend(g["subheaders"])

    # Write each item row
    for r_idx, row_dict in enumerate(all_item_rows, start=start_data_row):
        # Map existing keys to new layout values
        values_by_subheader = {
            # Source
            "New or Copied from Study": "",
            # Form
            "Form Label": gv(row_dict, "CTDM Optional, if blank CDP to propose"),
            "Form Name (provided by SDTM Programmer, if SDTM linked form)": gv(row_dict, "Input needed from SDTM"),
            # Item Group
            "Item Group (if only one on form, recommend same as Form Label)": gv(row_dict, "CDAI input needed"),
            "Item group Repeating": gv(row_dict, "Unnamed: 4"),
            "Repeat Maximum, if known, else default =50": gv(row_dict, "Unnamed: 5"),
            "Display format of repeating item group (Grid, read only, form)": gv(row_dict, "Unnamed: 6"),
            "Default Data in repeating item group": gv(row_dict, "Unnamed: 7"),
            # Item
            "Item Order": gv(row_dict, "Unnamed: 8"),
            "Item Label": gv(row_dict, "Unnamed: 9"),
            "Item Name (provided by SDTM Programmer, if SDTM linked item)": gv(row_dict, "Unnamed: 10"),
            # Progressive Display
            "Progressively displayed?": gv(row_dict, "Unnamed: 12"),
            "Controlling item (item triggering it, if yes, describe item below)": gv(row_dict, "Unnamed: 13"),
            "Controlling item value": gv(row_dict, "Unnamed: 14"),
            # Data Type
            "Data type": gv(row_dict, "Unnamed: 16"),
            "If text or number, Field Length": gv(row_dict, "Unnamed: 17"),
            "If number, Precision (decimal places)": gv(row_dict, "Unnamed: 18"),
            # Codelist
            "Codelist – Choice Labels (if binary, can use Goodlist Table)": gv(row_dict, "Unnamed: 19"),
            "Codelist Name (provided by SDTM Programmer)": gv(row_dict, "Unnamed: 20"),
            "Choice Code (provided by SDTM Programmer)": gv(row_dict, "Unnamed: 21"),
            "Codelist Control Type": gv(row_dict, "Unnamed: 22"),
            # System Queries
            "If number, Range: Min Value / Max Value": gv(row_dict, "Unnamed: 23"),
            "Date: Query Future Date": gv(row_dict, "Unnamed: 24"),
            "Required": gv(row_dict, "Unnamed: 25"),
            "If Required, Open Query when intentionally left blank (form/item)": gv(row_dict, "Unnamed: 26"),
            # Notes
            "Notes": gv(row_dict, "Unnamed: 27"),
        }

        for c_idx, header in enumerate(ordered_subheaders, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=values_by_subheader.get(header, ""))
            cell.alignment = left_top
            cell.border = thin_border

    # Apply borders to header rows and auto column widths
    for r in range(1, ws.max_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            # Add borders if not already set for headers
            if r <= 3:
                cell.border = thin_border

    # Auto width approximation
    for col in range(1, total_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            try:
                length = len(str(val))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(12, min(60, max_len + 2))

    # Save
    wb.save(output_csv_path)
    print(f"\n✅ SUCCESS! Created Study Specific Forms Excel: {output_csv_path} with {len(all_item_rows)} item rows.")
    print("✅ Header layout: 4 fixed CTDM rows with grouped headers applied.")


def write_study_specific_forms_stream(
    items_rows: list,
    workbook,
    sheet_name: str = "Study Specific Forms",
) -> None:
    """
    Stream-write Study Specific Forms sheet into an existing XlsxWriter workbook.
    Expects rows already mapped to the grouped header layout (same order as in
    process_clinical_forms Excel output). This is used by the combined PTD streaming mode.
    """
    # Define grouped headers and colors
    groups = [
        {"name": "Source", "subheaders": ["New or Copied from Study"], "color": "E7E6E6"},
        {"name": "Form", "subheaders": [
            "Form Label",
            "Form Name (provided by SDTM Programmer, if SDTM linked form)",
        ], "color": "C6EFCE"},
        {"name": "Item Group", "subheaders": [
            "Item Group (if only one on form, recommend same as Form Label)",
            "Item group Repeating",
            "Repeat Maximum, if known, else default =50",
            "Display format of repeating item group (Grid, read only, form)",
            "Default Data in repeating item group",
        ], "color": "B3E5FC"},
        {"name": "Item", "subheaders": [
            "Item Order",
            "Item Label",
            "Item Name (provided by SDTM Programmer, if SDTM linked item)",
        ], "color": "FFD7A8"},
        {"name": "Progressive Display", "subheaders": [
            "Progressively displayed?",
            "Controlling item (item triggering it, if yes, describe item below)",
            "Controlling item value",
        ], "color": "B3E5FC"},
        {"name": "Data Type", "subheaders": [
            "Data type",
            "If text or number, Field Length",
            "If number, Precision (decimal places)",
        ], "color": "FFF9C4"},
        {"name": "Codelist", "subheaders": [
            "Codelist – Choice Labels (if binary, can use Goodlist Table)",
            "Codelist Name (provided by SDTM Programmer)",
            "Choice Code (provided by SDTM Programmer)",
            "Codelist Control Type",
        ], "color": "E2F0D9"},
        {"name": "System Queries", "subheaders": [
            "If number, Range: Min Value / Max Value",
            "Date: Query Future Date",
            "Required",
            "If Required, Open Query when intentionally left blank (form/item)",
        ], "color": "F8CBAD"},
        {"name": "Notes", "subheaders": ["Notes"], "color": "E6B8AF"},
    ]

    ws = workbook.add_worksheet(sheet_name)

    header_font = workbook.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    subheader_font = workbook.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    ctdm_fill = workbook.add_format({'bg_color': '#F5F5F5', 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    left_top = workbook.add_format({'align': 'left', 'valign': 'top', 'text_wrap': True, 'border': 1})
    center = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})

    # Row 1: CTDM meta labels in A1:D1
    ctdm_titles = [
        "CTDM to fill in",
        "CTDM Optional, if blank CDP to propose",
        "Input needed from SDTM",
        "CDAI input needed",
    ]
    for idx, title in enumerate(ctdm_titles):
        ws.write(0, idx, title, ctdm_fill)

    # Row 2: Group names merged across subheaders
    col_start = 0
    for group in groups:
        width = len(group["subheaders"])
        start_col = col_start
        end_col = col_start + width - 1
        group_fill = workbook.add_format({'bg_color': f"#{group['color']}", 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})
        ws.merge_range(1, start_col, 1, end_col, group["name"], group_fill)
        for i in range(start_col, end_col + 1):
            ws.write(1, i, group["name"], group_fill)
        col_start += width

    # Row 3: Subheaders
    col_start = 0
    group_fills = []
    for group in groups:
        width = len(group["subheaders"])
        group_fill = workbook.add_format({'bg_color': f"#{group['color']}", 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})
        for i, sub in enumerate(group["subheaders"]):
            ws.write(2, col_start + i, sub, group_fill)
        group_fills.append(group_fill)
        col_start += width

    # Data rows start at row 4 (index 3)
    start_row = 3
    for r_offset, row_values in enumerate(items_rows):
        row_idx = start_row + r_offset
        for c_idx, val in enumerate(row_values):
            fmt = left_top if c_idx in {1, 2, 3} else center
            ws.write(row_idx, c_idx, val, fmt)


def get_groups_spec():
    return [
        {
            "name": "Source",
            "subheaders": [
                "New or Copied from Study",
            ],
            "color": "E7E6E6",
        },
        {
            "name": "Form",
            "subheaders": [
                "Form Label",
                "Form Name (provided by SDTM Programmer, if SDTM linked form)",
            ],
            "color": "C6EFCE",
        },
        {
            "name": "Item Group",
            "subheaders": [
                "Item Group (if only one on form, recommend same as Form Label)",
                "Item group Repeating",
                "Repeat Maximum, if known, else default =50",
                "Display format of repeating item group (Grid, read only, form)",
                "Default Data in repeating item group",
            ],
            "color": "B3E5FC",
        },
        {
            "name": "Item",
            "subheaders": [
                "Item Order",
                "Item Label",
                "Item Name (provided by SDTM Programmer, if SDTM linked item)",
            ],
            "color": "FFD7A8",
        },
        {
            "name": "Progressive Display",
            "subheaders": [
                "Progressively displayed?",
                "Controlling item (item triggering it, if yes, describe item below)",
                "Controlling item value",
            ],
            "color": "B3E5FC",
        },
        {
            "name": "Data Type",
            "subheaders": [
                "Data type",
                "If text or number, Field Length",
                "If number, Precision (decimal places)",
            ],
            "color": "FFF9C4",
        },
        {
            "name": "Codelist",
            "subheaders": [
                "Codelist – Choice Labels (if binary, can use Goodlist Table)",
                "Codelist Name (provided by SDTM Programmer)",
                "Choice Code (provided by SDTM Programmer)",
                "Codelist Control Type",
            ],
            "color": "E2F0D9",
        },
        {
            "name": "System Queries",
            "subheaders": [
                "If number, Range: Min Value / Max Value",
                "Date: Query Future Date",
                "Required",
                "If Required, Open Query when intentionally left blank (form/item)",
            ],
            "color": "F8CBAD",
        },
        {
            "name": "Notes",
            "subheaders": [
                "Notes",
            ],
            "color": "E6B8AF",
        },
    ]


def prepare_study_specific_forms_rows(
    json_file_path: str,
    config_path: str = "./config/config_study_specific_forms.json",
):
    """
    Build the Study Specific Forms rows (values only) in the same order as the
    grouped header layout. This avoids building an in-memory workbook and enables
    streaming write.
    Returns: list of lists (each inner list corresponds to ordered subheaders).
    """
    global CONFIG
    CONFIG = load_config(config_path)

    with open(json_file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    extracted_forms = extract_forms_cleaned(data)

    all_item_rows = []
    for form in extracted_forms:
        items = extract_items_from_form(form['Form_Node'])
        if not items:
            items.append({"Item Name": "", "Option_TD_Node": None, "Item Group": ""})
        items = assign_item_order(items)
        item_group_counts, repeating_groups = analyze_item_groups_per_form(items)

        for item in items:
            option_node = item.get("Option_TD_Node")
            item_name = item['Item Name']
            item_group_value = item.get("Item Group", "") or 'NaN'
            item_group_repeating_flag = get_item_group_repeating_flag(item_group_value, repeating_groups)
            repeat_maximum = get_repeat_maximum(item_group_value, item_group_repeating_flag, item_group_counts)
            item_order = item.get('Item_Order', 1)

            codelist_content = get_all_lbody_values(option_node)
            data_type = determine_data_type(option_node, codelist_content)
            field_length = calculate_field_length(codelist_content) if data_type in ["Text", "Label"] else ""
            precision = calculate_precision(codelist_content) if data_type == "Label" else ""
            number_range = extract_number_range(codelist_content) if data_type == "Label" else ""
            query_future_date = check_query_future_date(data_type)
            is_required = check_required_field(item_name)

            row_dict = {
                'CTDM Optional, if blank CDP to propose': form['Form Label'],
                'Input needed from SDTM': form['Form Name'],
                'CDAI input needed': item_group_value,
                'Unnamed: 4': item_group_repeating_flag,
                'Unnamed: 5': repeat_maximum,
                'Unnamed: 8': item_order,
                'Unnamed: 9': item_name,
                'Unnamed: 10': "",
                'Unnamed: 16': data_type,
                'Unnamed: 17': field_length,
                'Unnamed: 18': precision,
                'Unnamed: 19': codelist_content,
                'Unnamed: 23': number_range,
                'Unnamed: 24': query_future_date,
                'Unnamed: 25': is_required,
                'Unnamed: 26': "Form,Item" if is_required == "Y" else "",
                'Unnamed: 27': "",
            }
            all_item_rows.append(row_dict)

    # Map row_dict to ordered subheaders
    groups = get_groups_spec()
    ordered_subheaders = []
    for g in groups:
        ordered_subheaders.extend(g["subheaders"])

    def gv(d, key, default=""):
        return d.get(key, default)

    rows: list = []
    for row_dict in all_item_rows:
        values_by_subheader = {
            # Source
            "New or Copied from Study": "",
            # Form
            "Form Label": gv(row_dict, "CTDM Optional, if blank CDP to propose"),
            "Form Name (provided by SDTM Programmer, if SDTM linked form)": gv(row_dict, "Input needed from SDTM"),
            # Item Group
            "Item Group (if only one on form, recommend same as Form Label)": gv(row_dict, "CDAI input needed"),
            "Item group Repeating": gv(row_dict, "Unnamed: 4"),
            "Repeat Maximum, if known, else default =50": gv(row_dict, "Unnamed: 5"),
            "Display format of repeating item group (Grid, read only, form)": gv(row_dict, "Unnamed: 6"),
            "Default Data in repeating item group": gv(row_dict, "Unnamed: 7"),
            # Item
            "Item Order": gv(row_dict, "Unnamed: 8"),
            "Item Label": gv(row_dict, "Unnamed: 9"),
            "Item Name (provided by SDTM Programmer, if SDTM linked item)": gv(row_dict, "Unnamed: 10"),
            # Progressive Display
            "Progressively displayed?": gv(row_dict, "Unnamed: 12"),
            "Controlling item (item triggering it, if yes, describe item below)": gv(row_dict, "Unnamed: 13"),
            "Controlling item value": gv(row_dict, "Unnamed: 14"),
            # Data Type
            "Data type": gv(row_dict, "Unnamed: 16"),
            "If text or number, Field Length": gv(row_dict, "Unnamed: 17"),
            "If number, Precision (decimal places)": gv(row_dict, "Unnamed: 18"),
            # Codelist
            "Codelist – Choice Labels (if binary, can use Goodlist Table)": gv(row_dict, "Unnamed: 19"),
            "Codelist Name (provided by SDTM Programmer)": gv(row_dict, "Unnamed: 20"),
            "Choice Code (provided by SDTM Programmer)": gv(row_dict, "Unnamed: 21"),
            "Codelist Control Type": gv(row_dict, "Unnamed: 22"),
            # System Queries
            "If number, Range: Min Value / Max Value": gv(row_dict, "Unnamed: 23"),
            "Date: Query Future Date": gv(row_dict, "Unnamed: 24"),
            "Required": gv(row_dict, "Unnamed: 25"),
            "If Required, Open Query when intentionally left blank (form/item)": gv(row_dict, "Unnamed: 26"),
            # Notes
            "Notes": gv(row_dict, "Unnamed: 27"),
        }
        rows.append([values_by_subheader.get(h, "") for h in ordered_subheaders])

    return rows



if __name__ == "__main__":

    import sys

    json_file = sys.argv[1] if len(sys.argv) > 1 else None
    if not json_file:
        print("Please provide JSON input file path as argument.")
        sys.exit(1)

    template_file = "template.xlsx"
    output_file = "Study_Specific_Form.xlsx"

    try:
        print("=" * 80)
        print("CLINICAL FORMS PROCESSING - WITH SEQUENTIAL ITEM ORDER (1, 2, 3...)")
        print("=" * 80)
        process_clinical_forms(json_file, template_csv_path="template.xlsx", output_csv_path="Study_Specific_Form.xlsx")
        print("\n🎯 PROCESSING COMPLETE!")
        print("✅ Key features of this version:")
        print("   1. ✅ Correctly handles items in <TH> + <TD> row structures.")
        print("   2. ✅ Data Type and Codelist values are item-specific.")
        print("   3. ✅ 'Item group Repeating' column filled based on occurrence count.")
        print("   4. ✅ 'Repeat Maximum' column set to count or default 50.")
        print("   5. 🔥 NEW: 'Item Order' filled sequentially (1, 2, 3...) per form.")
        print("   6. ✅ Generates one row per unique item found in a form.")

    except Exception as e:
        print(f"❌ An error occurred: {e}")
        import traceback

        traceback.print_exc()