#!/usr/bin/env python3
"""
Generate a combined PTD workbook with two sheets:
 - Sheet 1: Schedule Grid (from protocol + eCRF)
 - Sheet 2: Study Specific Forms (from eCRF)

The computation logic is reused from existing modules and scripts; this file
organizes them into a single configurable CLI entrypoint and merges outputs.

Modifications:
 - Added --template argument to load an existing template workbook
 - Generate Schedule Grid and Study Specific Forms into temporary files
 - Replace sheets named "Schedule Grid" and "Study Specific Forms" in the template
   (including styles, merges, and dimensions), preserve other sheets, and save to --out
 - Removed the old append/merge flow that built a new workbook from scratch
"""

import os
import sys
import json
import logging
import argparse
from typing import Dict, Any, Optional, List
from pathlib import Path
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Reuse existing modules for schedule grid pipeline
from .modules.form_extractor import extract_forms
from .modules.soa_parser import parse_soa
from .modules.common_matrix import merge_common_matrix
from .modules.event_grouping import group_events
from .modules.schedule_layout import generate_schedule_grid as build_schedule_grid_file
from .modules.schedule_layout import generate_schedule_grid_stream
from .Final_study_specific_form import prepare_study_specific_forms_rows, write_study_specific_forms_stream


def load_json(file_path: str) -> Dict[str, Any]:
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_config(config_path: str) -> Dict[str, Any]:
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        logging.warning(f"Config file not found: {config_path}; using defaults")
        return {}
    except json.JSONDecodeError as e:
        logging.warning(f"Invalid JSON in config {config_path}: {e}; using defaults")
        return {}


def setup_logging(level: str = "INFO") -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper()),
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler("ptd_generation.log")
        ]
    )
def _col_to_letter(n: int) -> str:
    """1-based column index to Excel letters."""
    result = []
    while n:
        n, rem = divmod(n - 1, 26)
        result.append(chr(65 + rem))
    return ''.join(reversed(result))


def _write_minimal_sheet_xml_from_rows(rows_iter, out_path: str) -> None:
    """
    Write a minimal worksheet XML with inline strings and numeric values only.
    Avoids sharedStrings/styles to keep it self-contained and low memory.
    rows_iter yields lists/tuples of cell values per row (1-based rows assumed sequential).
    """
    import html

    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
        f.write(f'<worksheet xmlns="{ns}">')
        f.write('<sheetData>')
        row_idx = 0
        for row in rows_iter:
            row_idx += 1
            if row is None:
                continue
            f.write(f'<row r="{row_idx}">')
            for col_idx, val in enumerate(row, start=1):
                if val is None or val == '':
                    continue
                cell_ref = f'{_col_to_letter(col_idx)}{row_idx}'
                # Numbers vs strings
                is_number = isinstance(val, (int, float)) and not (isinstance(val, float) and (val != val or val in (float("inf"), float("-inf"))))
                if is_number:
                    f.write(f'<c r="{cell_ref}"><v>{val}</v></c>')
                else:
                    text = html.escape(str(val))
                    f.write(f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>')
            f.write('</row>')
        f.write('</sheetData>')
        f.write('</worksheet>')


def _iter_rows_from_xlsx_sheet(xlsx_path: str):
    """Yield rows as lists from the first worksheet of an XLSX file in read-only mode."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        try:
            wb.close()
        except Exception:
            pass



def ensure_output_dir(output_path: str) -> None:
    out_dir = os.path.dirname(output_path)
    if out_dir:
        Path(out_dir).mkdir(parents=True, exist_ok=True)


def run_schedule_grid_pipeline(
    protocol_json: str,
    ecrf_json: str,
    final_output_xlsx: str,
    config_dir: str,
    for_stream: bool = False,
) -> Any:
    """
    Reuse the existing 5-stage pipeline to produce inputs and/or the schedule grid.

    When for_stream=False (default):
        - Produces the schedule grid Excel at final_output_xlsx and returns its absolute path.
    When for_stream=True:
        - Produces intermediates needed for streaming (visits_xlsx, matrix_csv) and returns
          a dict { 'visits_xlsx', 'matrix_csv', 'temp_dir' }. The final schedule grid file
          is not built to save time/memory.
    """
    config_files = {
        'form_extractor': 'config_form_extractor.json',
        'soa_parser': 'config_soa_parser.json',
        'common_matrix': 'config_common_matrix.json',
        'event_grouping': 'config_event_grouping.json',
        'schedule_layout': 'config_schedule_layout.json'
    }

    configs: Dict[str, Any] = {}
    for key, filename in config_files.items():
        configs[key] = load_config(os.path.join(config_dir, filename))

    temp_dir = tempfile.mkdtemp(prefix="ptd_intermediate_")
    intermediates: List[str] = []
    try:
        forms_csv = os.path.join(temp_dir, "extracted_forms.csv")
        extract_forms(ecrf_json=ecrf_json, output_csv=forms_csv, config=configs.get('form_extractor', {}))
        intermediates.append(forms_csv)

        schedule_csv = os.path.join(temp_dir, "schedule.csv")
        parse_soa(protocol_json=protocol_json, output_csv=schedule_csv, config=configs.get('soa_parser', {}))
        intermediates.append(schedule_csv)

        matrix_csv = os.path.join(temp_dir, "soa_matrix.csv")
        merge_common_matrix(ecrf_csv=forms_csv, schedule_csv=schedule_csv, output_csv=matrix_csv, config=configs.get('common_matrix', {}))
        intermediates.append(matrix_csv)

        visits_xlsx = os.path.join(temp_dir, "visits_with_groups.xlsx")
        group_events(protocol_json=protocol_json, output_xlsx=visits_xlsx, config=configs.get('event_grouping', {}))
        intermediates.append(visits_xlsx)

        if for_stream:
            # Return the two inputs required for streaming writer; do NOT build final workbook
            return {
                'visits_xlsx': visits_xlsx,
                'matrix_csv': matrix_csv,
                'temp_dir': temp_dir,
                'intermediates': list(intermediates),
            }
        else:
            ensure_output_dir(final_output_xlsx)
            build_schedule_grid_file(
                visits_xlsx=visits_xlsx,
                forms_csv=matrix_csv,
                output_xlsx=final_output_xlsx,
                config=configs.get('schedule_layout', {}),
            )
            return os.path.abspath(final_output_xlsx)
    finally:
        # Best-effort cleanup; keep intermediates when for_stream
        if not for_stream:
            for p in intermediates:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception:
                    pass
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


def generate_study_specific_forms_xlsx(ecrf_json: str) -> str:
    """
    Reuse logic from Final_study_specific_form.py by invoking its processing function to
    produce an Excel file. Returns the path to the generated temp Excel.
    """
    # Import here to avoid executing module-level code unless needed
    import importlib.util

    module_path = os.path.join(os.path.dirname(__file__), 'Final_study_specific_form.py')
    spec = importlib.util.spec_from_file_location("Final_study_specific_form", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Unable to load Final_study_specific_form.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    temp_dir = tempfile.mkdtemp(prefix="ptd_forms_")
    output_xlsx = os.path.join(temp_dir, "study_specific_forms.xlsx")

    # The script's API function writes the Excel; keep its computation logic intact
    config_rules = os.path.join(os.path.dirname(__file__), 'config', 'config_study_specific_forms.json')
    # Avoid hardcoded/unnecessary template path; rely on the module's internal template
    mod.process_clinical_forms(ecrf_json, output_csv_path=output_xlsx, config_path=config_rules)
    return output_xlsx


## Removed: unused header renaming/ordering helper.


def _copy_worksheet_contents(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    """Copy values, styles, merged cells, and dimensions from src_ws to dest_ws."""
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if getattr(dim, 'width', None):
            dest_ws.column_dimensions[col_letter].width = dim.width

    # Copy row heights
    for idx, dim in src_ws.row_dimensions.items():
        if getattr(dim, 'height', None):
            dest_ws.row_dimensions[idx].height = dim.height

    # Copy merged ranges first (structure)
    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))

    # Copy cell contents and styles
    for row in src_ws.iter_rows():
        for cell in row:
            # Skip non-top-left merged cells; the range has already been created
            if isinstance(cell, MergedCell):
                continue
            dcell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                if cell.font:
                    dcell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color,
                    )
                if cell.alignment:
                    dcell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent,
                    )
                if cell.fill and cell.fill.fill_type:
                    dcell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color,
                    )
                if cell.border:
                    left = cell.border.left
                    right = cell.border.right
                    top = cell.border.top
                    bottom = cell.border.bottom
                    dcell.border = Border(
                        left=Side(style=left.style, color=left.color),
                        right=Side(style=right.style, color=right.color),
                        top=Side(style=top.style, color=top.color),
                        bottom=Side(style=bottom.style, color=bottom.color),
                    )
                if cell.number_format:
                    dcell.number_format = cell.number_format


def _copy_worksheet_values_only(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    """
    Fast path: copy only cell values row-by-row using append.
    Skips styles, merges, widths, heights for speed.
    """
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))

    # Preserve merged cell structure if available (ReadOnlyWorksheet may not expose it)
    try:
        merged_ranges = None
        merged_container = getattr(src_ws, "merged_cells", None)
        if merged_container is not None:
            merged_ranges = getattr(merged_container, "ranges", None)
        if not merged_ranges:
            # Fallback for older openpyxl
            merged_ranges = getattr(src_ws, "merged_cell_ranges", None)
        if merged_ranges:
            for merged_range in merged_ranges:
                try:
                    dest_ws.merge_cells(str(merged_range))
                except Exception:
                    pass
    except Exception:
        # If unavailable in read-only mode, skip merges for speed/simplicity
        pass


def _copy_header_styles(src_ws: Worksheet, dest_ws: Worksheet, max_header_rows: int) -> None:
    """
    Copy only styles (font, alignment, fill, border, number_format) for the first
    max_header_rows to preserve header/group colouring while keeping operation fast.
    """
    if max_header_rows <= 0:
        return
    rows_to_copy = min(max_header_rows, src_ws.max_row)
    for r in range(1, rows_to_copy + 1):
        for cell in src_ws[r]:
            try:
                dcell = dest_ws.cell(row=cell.row, column=cell.column)
                if cell.has_style:
                    if cell.font:
                        dcell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                        )
                    if cell.alignment:
                        dcell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )
                    if cell.fill and cell.fill.fill_type:
                        dcell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                        )
                    if cell.border:
                        left = cell.border.left
                        right = cell.border.right
                        top = cell.border.top
                        bottom = cell.border.bottom
                        dcell.border = Border(
                            left=Side(style=left.style, color=left.color),
                            right=Side(style=right.style, color=right.color),
                            top=Side(style=top.style, color=top.color),
                            bottom=Side(style=bottom.style, color=bottom.color),
                        )
                    if cell.number_format:
                        dcell.number_format = cell.number_format
            except Exception:
                # Best-effort; continue
                pass


def replace_sheets_in_template(
    template_xlsx: str,
    schedule_xlsx: str,
    forms_xlsx: str,
    out_xlsx: str,
    schedule_sheet_name: str = "Schedule Grid",
    forms_sheet_name: str = "Study Specific Forms",
    fast: bool = False,
) -> str:
    """
    Load the template workbook, remove existing target sheets if present, copy
    the generated schedule and forms worksheets (including styles, merges, and
    dimensions) into the template, preserve all other sheets, and save to out_xlsx.
    Returns the absolute path to the saved workbook.
    """
    ensure_output_dir(out_xlsx)

    # Use read_only to reduce memory; for fast path we copy only values+header styles
    wb_template = load_workbook(template_xlsx, read_only=False)
    wb_schedule = load_workbook(schedule_xlsx, read_only=(fast is True), data_only=True)
    wb_forms = load_workbook(forms_xlsx, read_only=(fast is True), data_only=True)

    try:
        # Determine insertion indices to preserve original order if sheets existed
        schedule_index = None
        forms_index = None
        if schedule_sheet_name in wb_template.sheetnames:
            schedule_index = wb_template.sheetnames.index(schedule_sheet_name)
            wb_template.remove(wb_template[schedule_sheet_name])
        if forms_sheet_name in wb_template.sheetnames:
            forms_index = wb_template.sheetnames.index(forms_sheet_name)
            wb_template.remove(wb_template[forms_sheet_name])

        # Create destination sheets at recorded positions (or append if None)
        if schedule_index is not None:
            dest_schedule = wb_template.create_sheet(title=schedule_sheet_name, index=schedule_index)
        else:
            dest_schedule = wb_template.create_sheet(title=schedule_sheet_name)
        if forms_index is not None:
            dest_forms = wb_template.create_sheet(title=forms_sheet_name, index=forms_index)
        else:
            dest_forms = wb_template.create_sheet(title=forms_sheet_name)

        # Source sheets (first worksheet in each generated file)
        src_schedule: Worksheet = wb_schedule.worksheets[0]
        src_forms: Worksheet = wb_forms.worksheets[0]

        # Copy contents
        if fast:
            _copy_worksheet_values_only(src_schedule, dest_schedule)
            _copy_worksheet_values_only(src_forms, dest_forms)
            # Preserve header styles (keep column/group colouring)
            _copy_header_styles(src_schedule, dest_schedule, max_header_rows=5)
            _copy_header_styles(src_forms, dest_forms, max_header_rows=3)
        else:
            _copy_worksheet_contents(src_schedule, dest_schedule)
            _copy_worksheet_contents(src_forms, dest_forms)

        wb_template.save(out_xlsx)
        return os.path.abspath(out_xlsx)
    finally:
        try:
            wb_schedule.close()
        except Exception:
            pass
        try:
            wb_forms.close()
        except Exception:
            pass
        try:
            wb_template.close()
        except Exception:
            pass


def auto_format_sheet(sheet: Worksheet, header_rows: int = 1, skip_fill_rows=None) -> None:
    """Auto-fit columns, style headers (one or more rows), and apply borders.

    Preserves any existing header fills and avoids filling Row 1 beyond column D.
    """
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style header rows
    header_rows = max(1, min(header_rows, sheet.max_row))
    skip_fill_rows = skip_fill_rows or set()
    for r in range(1, header_rows + 1):
        for cell in sheet[r]:
            cell.font = header_font
            cell.alignment = center_align
            # Preserve pre-existing fills (do not override group colors)
            try:
                fill_type = getattr(cell.fill, 'fill_type', None)
            except Exception:
                fill_type = None
            # Do not apply header fill to Row 1 columns beyond D
            col_idx = cell.column if hasattr(cell, 'column') else cell.col_idx
            beyond_ctdm = (r == 1 and col_idx and col_idx > 4)
            if not fill_type and r not in skip_fill_rows and not beyond_ctdm:
                cell.fill = header_fill
            cell.border = thin_border

    # Style all data rows
    for row in sheet.iter_rows(min_row=header_rows + 1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        first_cell = next(iter(col))
        column_letter = get_column_letter(first_cell.column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        sheet.column_dimensions[column_letter].width = max(10, min(80, max_length + 3))


def finalize_formatting(output_path: str, forms_sheet_name: str = "Study Specific Forms") -> None:
    """
    Apply formatting only to the study-specific forms sheet. The schedule grid
    formatting produced by its generator is preserved as-is.
    """
    wb = load_workbook(output_path)
    if forms_sheet_name in wb.sheetnames:
        ws = wb[forms_sheet_name]
        # Keep the 3 fixed header rows (Row 1 CTDM, Row 2 merged groups, Row 3 subheaders)
        auto_format_sheet(ws, header_rows=3, skip_fill_rows={})
    wb.save(output_path)


def surgery_replace_sheets_inplace(
    template_xlsx: str,
    schedule_sheet_name: str,
    forms_sheet_name: str,
    schedule_rows_path: str,
    forms_rows_iter,
) -> str:
    """
    Perform a low-memory zip-level transplant: replace only the two target sheet XMLs
    in the template, leave all other parts intact. The replacement sheet XMLs are
    minimal (values only, inline strings, no external styles/sharedStrings).
    Returns the absolute path to the modified template (in-place).
    """
    import zipfile
    import tempfile
    import shutil
    import xml.etree.ElementTree as ET
    from pathlib import Path

    tmp_dir = tempfile.mkdtemp(prefix='ptd_surgery_')
    extract_dir = os.path.join(tmp_dir, 'extract')
    os.makedirs(extract_dir, exist_ok=True)

    # Extract template
    with zipfile.ZipFile(template_xlsx, 'r') as zin:
        zin.extractall(extract_dir)

    # Parse workbook and rels to find sheet targets
    wb_xml = os.path.join(extract_dir, 'xl', 'workbook.xml')
    rels_xml = os.path.join(extract_dir, 'xl', '_rels', 'workbook.xml.rels')
    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main', 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
    tree = ET.parse(wb_xml)
    root = tree.getroot()
    name_to_rid = {}
    for sheet in root.findall('ns:sheets/ns:sheet', ns):
        name = sheet.attrib.get('name')
        rid = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        if name and rid:
            name_to_rid[name] = rid
    if schedule_sheet_name not in name_to_rid or forms_sheet_name not in name_to_rid:
        raise RuntimeError('Target sheet names not found in template workbook.')

    rels_tree = ET.parse(rels_xml)
    rels_root = rels_tree.getroot()
    rid_to_target = {}
    for rel in rels_root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
        rid_to_target[rel.attrib.get('Id')] = rel.attrib.get('Target')

    def target_to_path(target: str) -> str:
        # Normalize odd targets like '/xl/worksheets/sheet1.xml' or '/worksheets/sheet1.xml'
        t = (target or '').replace('\\', '/').lstrip('/')
        if t.startswith('xl/'):
            t = t[3:]
        # Usually ends as 'worksheets/sheet1.xml'
        full = os.path.join(extract_dir, 'xl', t)
        os.makedirs(os.path.dirname(full), exist_ok=True)
        return full

    sched_target = rid_to_target[name_to_rid[schedule_sheet_name]]
    forms_target = rid_to_target[name_to_rid[forms_sheet_name]]
    sched_path = target_to_path(sched_target)
    forms_path = target_to_path(forms_target)

    # Build replacement xmls
    # 1) Schedule: source is an XLSX on disk; stream rows and write xml
    _write_minimal_sheet_xml_from_rows(_iter_rows_from_xlsx_sheet(schedule_rows_path), sched_path)

    # 2) Forms: rows iterator provided
    _write_minimal_sheet_xml_from_rows(forms_rows_iter, forms_path)

    # Repack zip (overwrite original)
    new_zip = os.path.join(tmp_dir, 'out.xlsx')
    with zipfile.ZipFile(new_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for folder, _, files in os.walk(extract_dir):
            for filename in files:
                full_path = os.path.join(folder, filename)
                arcname = os.path.relpath(full_path, extract_dir)
                zout.write(full_path, arcname)

    # Replace original
    shutil.copyfile(new_zip, template_xlsx)
    shutil.rmtree(tmp_dir, ignore_errors=True)
    return os.path.abspath(template_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate PTD Excel with Schedule Grid and Study Specific Forms"
    )
    parser.add_argument("--ecrf", required=True, help="Path to hierarchical_output_final_ecrf.json")
    parser.add_argument("--protocol", required=True, help="Path to hierarchical_output_final_protocol.json")
    parser.add_argument("--template", required=False, help="Path to template Excel (will be updated)")
    parser.add_argument("--out", required=False, help="Output Excel file path (e.g., ptd.xlsx). Omit when using --inplace")
    parser.add_argument("--inplace", action="store_true", help="Modify the template file in place (save over --template)")
    parser.add_argument("--fast", action="store_true", help="Fast mode: values-only copy, skip extra formatting")
    parser.add_argument("--stream", action="store_true", help="Stream directly to a new workbook using XlsxWriter (preserves formatting and minimizes memory)")
    parser.add_argument("--surgery", action="store_true", help="Low-RAM in-place surgery: replace only target sheet XMLs in the template")
    args = parser.parse_args()

    setup_logging("INFO")

    # Determine output path (in-place or new file)
    if args.stream:
        if not args.out:
            print("Error: --out is required for --stream mode", file=sys.stderr)
            return 2
        output_path = args.out
    else:
        if args.inplace:
            if not args.template:
                print("Error: --template is required for --inplace mode", file=sys.stderr)
                return 2
            output_path = args.template
            if args.out and os.path.abspath(args.out) != os.path.abspath(args.template):
                logging.warning("--inplace specified: ignoring --out and writing to template path")
        else:
            if not args.out:
                print("Error: --out is required unless --inplace is specified", file=sys.stderr)
                return 2
            output_path = args.out

    # Normalize output path extension and ensure folder
    if not output_path.lower().endswith(".xlsx"):
        output_path = os.path.splitext(output_path)[0] + ".xlsx"
    ensure_output_dir(output_path)

    # 1) Build schedule grid into a temp workbook
    schedule_tmp_dir = tempfile.mkdtemp(prefix="ptd_schedule_")
    schedule_tmp_xlsx = os.path.join(schedule_tmp_dir, "schedule_grid.xlsx")
    schedule_inputs = run_schedule_grid_pipeline(
        protocol_json=args.protocol,
        ecrf_json=args.ecrf,
        final_output_xlsx=schedule_tmp_xlsx,
        config_dir=os.path.join(os.path.dirname(__file__), "config"),
        for_stream=(args.stream or args.surgery),
    )

    # 2) Generate study specific forms to a temp file (only in non-stream mode)
    if not args.stream:
        forms_tmp_xlsx = generate_study_specific_forms_xlsx(args.ecrf)

    if args.stream:
        # Stream both sheets into a single workbook using XlsxWriter
        import xlsxwriter
        ensure_output_dir(output_path)
        workbook = xlsxwriter.Workbook(output_path, {
            'constant_memory': True,
            'strings_to_urls': False,
        })
        try:
            # Schedule Grid
            generate_schedule_grid_stream(
                visits_xlsx=schedule_inputs['visits_xlsx'],
                forms_csv=schedule_inputs['matrix_csv'],
                workbook=workbook,
                sheet_name="Schedule Grid",
                config=load_config(os.path.join(os.path.dirname(__file__), "config", "config_schedule_layout.json")),
            )

            # Study Specific Forms
            rows = prepare_study_specific_forms_rows(
                json_file_path=args.ecrf,
                config_path=os.path.join(os.path.dirname(__file__), 'config', 'config_study_specific_forms.json'),
            )
            write_study_specific_forms_stream(rows, workbook, sheet_name="Study Specific Forms")
        finally:
            workbook.close()
        final_path = output_path
        # Cleanup streaming intermediates created by run_schedule_grid_pipeline
        try:
            shutil.rmtree(schedule_inputs['temp_dir'])
        except Exception:
            pass
    elif args.surgery:
        # Build forms rows iterator without creating a big workbook
        rows = prepare_study_specific_forms_rows(
            json_file_path=args.ecrf,
            config_path=os.path.join(os.path.dirname(__file__), 'config', 'config_study_specific_forms.json'),
        )
        # Perform zip-level sheet transplant in-place
        final_path = surgery_replace_sheets_inplace(
            template_xlsx=output_path,
            schedule_sheet_name="Schedule Grid",
            forms_sheet_name="Study Specific Forms",
            schedule_rows_path=schedule_inputs['visits_xlsx'],
            forms_rows_iter=iter(rows),
        )
        # Clean intermediates
        try:
            shutil.rmtree(schedule_inputs['temp_dir'])
        except Exception:
            pass
    else:
        # 3) Replace sheets in the provided template and save to output
        if not args.template:
            print("Error: --template is required when not using --stream", file=sys.stderr)
            return 2
        final_path = replace_sheets_in_template(
            template_xlsx=args.template,
            schedule_xlsx=schedule_inputs,
            forms_xlsx=forms_tmp_xlsx,
            out_xlsx=output_path,
            fast=args.fast,
        )

        # 4) Finalize formatting on forms sheet (skip in fast mode)
        if not args.fast:
            finalize_formatting(final_path)

    # Cleanup temp dirs
    try:
        shutil.rmtree(schedule_tmp_dir)
    except Exception:
        pass
    if not args.stream:
        try:
            shutil.rmtree(os.path.dirname(forms_tmp_xlsx))
        except Exception:
            pass

    print(f"✅ Combined PTD file written successfully to: {final_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())