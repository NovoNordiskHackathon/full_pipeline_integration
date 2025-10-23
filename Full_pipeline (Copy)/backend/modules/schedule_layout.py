"""
Schedule Layout Module

Generates the final schedule grid layout from visit groups and forms data,
creating a comprehensive Excel output for clinical trial planning.
"""

import re
import logging
import pandas as pd
import math
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional


def make_event_name(group: str, label: str, idx: int, config: Dict[str, Any]) -> str:
    """Generate short event name from group and label."""
    event_mapping = config.get('event_name_mapping', {})
    
    g = str(group).strip().lower()
    s = str(label).strip().lower()
    
    # Check explicit overrides
    if "screen" in g:
        return event_mapping.get('screening', 'SCRN')
    if "random" in g:
        return event_mapping.get('random', 'RAND')
    if "rtsm" in g:
        return event_mapping.get('rtsm', 'RTSM')
    
    # Detect visit numbers from label
    visit_pattern = event_mapping.get('visit_pattern', 'V{number}')
    phone_pattern = event_mapping.get('phone_pattern', 'P{number}')
    
    m = re.search(r'\bV\s*?(\d+)\b', s) or re.search(r'\bVisit\s*?(\d+)\b', s) or re.search(r'\bP(\d+)\b', s)
    if m:
        if 'P' in m.group(0):
            return phone_pattern.format(number=m.group(1))
        else:
            return visit_pattern.format(number=m.group(1))
    
    # Fallback
    return f"V{idx + 1}"


def build_schedule_layout(visit_schedule_xlsx: str, forms_csv: str, output_xlsx: str, 
                         config: Dict[str, Any] = None) -> str:
    """Build the final PTD schedule grid Excel layout and save to output_xlsx."""
    if config is None:
        config = {}
    
    logging.info(f"Building schedule layout from {visit_schedule_xlsx} and {forms_csv}")
    
    try:
        df_visits = pd.read_excel(visit_schedule_xlsx, sheet_name=0)
        df_forms = pd.read_csv(forms_csv)
    except Exception as e:
        logging.error(f"Error loading input files: {e}")
        raise
    
    # Normalize column names
    df_visits.columns = [c.strip() for c in df_visits.columns]
    df_forms.columns = [c.strip() for c in df_forms.columns]
    
    # Derive visit info
    visit_groups = df_visits["Event Group"].astype(str).tolist()
    visit_labels = df_visits["Visit Name"].astype(str).tolist()
    n_visits = len(visit_labels)
    
    # Generate event names
    event_names = [make_event_name(visit_groups[i], visit_labels[i], i, config) 
                   for i in range(n_visits)]
    
    # Locate first Randomisation index
    rand_idx = None
    for i, g in enumerate(visit_groups):
        if "random" in str(g).lower():
            rand_idx = i
            break
    if rand_idx is None:
        rand_idx = 0
    
    # Get configuration
    left_columns = config.get('left_columns', ['Form Label', 'Form Name', 'Source'])
    extra_headers = config.get('extra_headers', [
        'Common Forms', 'N/A', 'Is Form Dynamic?', 'Form Dynamic Criteria',
        'Additional Programming Instructions'
    ])
    
    # ------------------ workbook setup ------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Final PTD"
    
    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # ------------------ HEADER ------------------
    # Left columns
    for i, lbl in enumerate(left_columns):
        for row in range(1, 4):
            cell = ws.cell(row=row, column=i + 1, value=lbl if row == 2 else None)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border
    
    # Insert Event Group/Label/Name column after Source
    col_after_source = len(left_columns) + 1
    
    ws.cell(row=1, column=col_after_source, value="Event Group:").font = bold
    ws.cell(row=1, column=col_after_source).alignment = center
    ws.cell(row=1, column=col_after_source).fill = header_fill
    ws.cell(row=1, column=col_after_source).border = border
    
    ws.cell(row=2, column=col_after_source, value="Event Label:").font = bold
    ws.cell(row=2, column=col_after_source).alignment = center
    ws.cell(row=2, column=col_after_source).fill = header_fill
    ws.cell(row=2, column=col_after_source).border = border
    
    ws.cell(row=3, column=col_after_source, value="Event Name:").font = bold
    ws.cell(row=3, column=col_after_source).alignment = center
    ws.cell(row=3, column=col_after_source).fill = header_fill
    ws.cell(row=3, column=col_after_source).border = border
    
    # RTSM column
    col_rtsm = col_after_source + 1
    for r in (1, 2, 3):
        ws.cell(row=r, column=col_rtsm, value="RTSM").font = bold
        ws.cell(row=r, column=col_rtsm).alignment = center
        ws.cell(row=r, column=col_rtsm).fill = header_fill
        ws.cell(row=r, column=col_rtsm).border = border
    
    # Visits start after RTSM
    col_start_visits = col_rtsm + 1
    
    # Row 2: Event Label (full names from xl)
    for j, vlabel in enumerate(visit_labels):
        c = col_start_visits + j
        if event_names[j] == "SCRN":
            event_label = "Screening"
        elif event_names[j] == "RAND":
            event_label = "Randomisation"
        elif "V" in event_names[j]:
            event_label = f"Visit {event_names[j][1:]}"
        elif "P" in event_names[j]:
            event_label = f"Phone Visit {event_names[j][1:]}"
        ws.cell(row=2, column=c, value=event_label).font = bold
        ws.cell(row=2, column=c).alignment = center
        ws.cell(row=2, column=c).fill = header_fill
        ws.cell(row=2, column=c).border = border
    
    # Row 3: Event Name (short codes)
    for j, ename in enumerate(event_names):
        c = col_start_visits + j
        ws.cell(row=3, column=c, value=ename).font = bold
        ws.cell(row=3, column=c).alignment = center
        ws.cell(row=3, column=c).fill = header_fill
        ws.cell(row=3, column=c).border = border
    
    # Row 1: Event Group (merged)
    cur_group, group_start_col = None, None
    for j, g in enumerate(visit_groups):
        c = col_start_visits + j
        if cur_group is None:
            cur_group, group_start_col = g, c
        if g != cur_group:
            ws.merge_cells(start_row=1, start_column=group_start_col, end_row=1, end_column=c - 1)
            ws.cell(row=1, column=group_start_col, value=cur_group).font = bold
            ws.cell(row=1, column=group_start_col).alignment = center
            ws.cell(row=1, column=group_start_col).fill = header_fill
            for cc in range(group_start_col, c):
                ws.cell(row=1, column=cc).border = border
            cur_group, group_start_col = g, c
    if group_start_col is not None:
        ws.merge_cells(start_row=1, start_column=group_start_col, end_row=1, end_column=col_start_visits + n_visits - 1)
        ws.cell(row=1, column=group_start_col, value=cur_group).font = bold
        ws.cell(row=1, column=group_start_col).alignment = center
        ws.cell(row=1, column=group_start_col).fill = header_fill
        for cc in range(group_start_col, col_start_visits + n_visits):
            ws.cell(row=1, column=cc).border = border
    
    # ------------------ Extra headers ------------------
    for idx, h in enumerate(extra_headers):
        c = col_start_visits + n_visits + idx
        ws.cell(row=1, column=c, value="").fill = header_fill
        ws.cell(row=1, column=c).border = border
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
        ws.cell(row=3, column=c, value="").fill = header_fill
        ws.cell(row=3, column=c).border = border
    
    # ------------------ BLOCKS: Visit Dynamics + Event Window ------------------
    cur_row = 4
    dynamic_rows = [
        "Visit Dynamics (If Y, then Event should appear based on triggering criteria)",
        "Triggering: Event",
        "Triggering: Form",
        "Triggering: Item = Response (if specific response expected, else leave to accept any entered result)"
    ]
    event_window_rows = [
        "Assign Visit Window",
        "Offset Type (Previous Event, Specific Event, or None)",
        "Offset Days (Planned Visit Date, as calculated from Offset Event)",
        "Day Range - Early",
        "Day Range - Late"
    ]
    sections = [("Visit Dynamic Properties", dynamic_rows),
                ("Event Window Configuration", event_window_rows)]
    
    for section_title, attrs in sections:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(left_columns))
        st_cell = ws.cell(row=cur_row, column=1, value=section_title)
        st_cell.font = bold
        st_cell.alignment = center
        st_cell.fill = grey_fill
        st_cell.border = border
        cur_row += 1
        
        for attr in attrs:
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(left_columns))
            lbl_cell = ws.cell(row=cur_row, column=1, value=attr)
            lbl_cell.font = bold
            lbl_cell.alignment = left_align
            lbl_cell.fill = grey_fill
            lbl_cell.border = border
            
            for j in range(n_visits):
                c = col_start_visits + j
                mapped_value = ""
                
                if attr.startswith("Visit Dynamics"):
                    eg = str(visit_groups[j]).lower()
                    if j >= rand_idx and ("end of treatment" not in eg and "end of study" not in eg):
                        mapped_value = "Y"
                
                elif attr.startswith("Triggering: Event"):
                    if event_names[j] == "RAND":
                        mapped_value = "SCRN"
                    elif event_names[j].startswith("V") and j > 0:
                        mapped_value = event_names[j - 1]
                    elif event_names[j].lower() == "follow-up":
                        mapped_value = "EOT"
                
                elif attr.startswith("Triggering: Form"):
                    mapped_value = ""
                    if event_names[j] == "RAND":
                        mapped_value = "ELIGIBILITY_CRITERIA"
                    elif j > rand_idx:
                        if "V" in visit_labels[j]:
                            if mapped_value == "":
                                mapped_value = "RANDOMISATION"
                                break
                    else:
                        mapped_value = ""
                
                elif attr.startswith("Assign Visit Window"):
                    mapped_value = "Y"
                
                elif attr.startswith("Offset Type") and "Offset Type" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Offset Type", "")
                elif attr.startswith("Offset Days") and "Offset Days" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Offset Days", "")
                elif attr.startswith("Day Range - Early") and "Day Range - Early" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Day Range - Early", "")
                elif attr.startswith("Day Range - Late") and "Day Range - Late" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Day Range - Late", "")
                
                if pd.isna(mapped_value):
                    mapped_value = ""
                if isinstance(mapped_value, float) and math.isclose(mapped_value, int(mapped_value)):
                    mapped_value = int(mapped_value)
                
                ws.cell(row=cur_row, column=c, value=mapped_value).alignment = center
                ws.cell(row=cur_row, column=c).border = border
            
            ws.cell(row=cur_row, column=col_rtsm, value="").border = border
            cur_row += 1
    
    # ------------------ FORMS TABLE ------------------
    forms_start_row = cur_row
    df_forms_filtered = df_forms.copy()
    row_cursor = forms_start_row
    
    # RTSM row
    ws.cell(row=row_cursor, column=1, value="RTSM").alignment = left_align
    ws.cell(row=row_cursor, column=2, value="RTSM").alignment = left_align
    ws.cell(row=row_cursor, column=3, value="Library").alignment = left_align
    ws.cell(row=row_cursor, column=col_rtsm, value="X").alignment = center
    for idx in range(len(extra_headers)):
        ws.cell(row=row_cursor, column=col_start_visits + n_visits + idx, value="").alignment = center
    row_cursor += 1
    
    # Forms from CSV
    for _, r in df_forms_filtered.iterrows():
        ws.cell(row=row_cursor, column=1, value=r.get('Form Label', '')).alignment = left_align
        ws.cell(row=row_cursor, column=2, value=r.get('Form Name', '')).alignment = left_align
        ws.cell(row=row_cursor, column=3, value=r.get('Source', '')).alignment = left_align
        ws.cell(row=row_cursor, column=col_rtsm, value="").alignment = center
        
        for j, vlabel in enumerate(visit_labels):
            c = col_start_visits + j
            val = ""
            if vlabel in r.index:
                val = r[vlabel]
            else:
                en = event_names[j]
                if en in r.index:
                    val = r[en]
            if pd.isna(val):
                val = ""
            if isinstance(val, float) and math.isclose(val, int(val)):
                val = int(val)
            ws.cell(row=row_cursor, column=c, value=val).alignment = center
        
        extra_vals = {
            "Is Form Dynamic?": r.get("Is Form Dynamic?", "") or r.get("Is Form Dynamic", "") or r.get("IsDynamic", ""),
            "Form Dynamic Criteria": r.get("Form Dynamic Criteria", "") or r.get("Form Dynamic Criteria ", "")
        }
        for idx, colname in enumerate(extra_headers):
            c = col_start_visits + n_visits + idx
            ws.cell(row=row_cursor, column=c, value=extra_vals.get(colname, "")).alignment = center
        row_cursor += 1
    
    # ------------------ formatting ------------------
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                lv = str(cell.value)
                max_len = max(max_len, len(lv))
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)
    
    ws.freeze_panes = ws.cell(row=forms_start_row, column=col_rtsm)
    
    wb.save(output_xlsx)
    logging.info(f"Schedule grid saved to {output_xlsx}")
    return output_xlsx


def generate_schedule_grid(visits_xlsx: str, forms_csv: str, output_xlsx: str, 
                          config: Dict[str, Any] = None) -> str:
    """
    Generate the final schedule grid from visit groups and forms data.
    
    Args:
        visits_xlsx: Path to visits with groups Excel file
        forms_csv: Path to forms matrix CSV file
        output_xlsx: Path to output Excel file
        config: Configuration dictionary
        
    Returns:
        Path to output Excel file
    """
    if config is None:
        config = {}
    
    logging.info(f"Generating schedule grid from {visits_xlsx} and {forms_csv}")
    
    try:
        output_path = build_schedule_layout(visits_xlsx, forms_csv, output_xlsx, config)
        return output_path
    except Exception as e:
        logging.error(f"Error generating schedule grid: {e}")
        raise


def generate_schedule_grid_stream(
    visits_xlsx: str,
    forms_csv: str,
    workbook,
    sheet_name: str = "Schedule Grid",
    config: Dict[str, Any] = None,
) -> None:
    """
    Stream-write the schedule grid directly into an existing XlsxWriter workbook.
    Preserves the same formatting semantics as build_schedule_layout, but writes
    row-by-row with constant memory.

    Args:
        visits_xlsx: Path to visits-with-groups Excel (first sheet used)
        forms_csv: Path to forms matrix CSV file
        workbook: An xlsxwriter.Workbook instance (opened with constant_memory)
        sheet_name: Name of the sheet to create
        config: Optional configuration dictionary
    """
    import pandas as pd  # Local import to avoid imposing dependency at import time
    import math

    if config is None:
        config = {}

    # Load inputs
    df_visits = pd.read_excel(visits_xlsx, sheet_name=0)
    # We'll stream forms CSV in chunks to avoid loading it fully into memory
    # when writing the large forms table.

    # Normalize columns
    df_visits.columns = [str(c).strip() for c in df_visits.columns]

    visit_groups = df_visits["Event Group"].astype(str).tolist()
    visit_labels = df_visits["Visit Name"].astype(str).tolist()
    num_visits = len(visit_labels)

    event_names = [
        make_event_name(visit_groups[i], visit_labels[i], i, config)
        for i in range(num_visits)
    ]

    # Locate first Randomisation index
    rand_idx = None
    for i, g in enumerate(visit_groups):
        if "random" in str(g).lower():
            rand_idx = i
            break
    if rand_idx is None:
        rand_idx = 0

    left_columns = config.get('left_columns', ['Form Label', 'Form Name', 'Source'])
    extra_headers = config.get('extra_headers', [
        'Common Forms', 'N/A', 'Is Form Dynamic?', 'Form Dynamic Criteria',
        'Additional Programming Instructions'
    ])

    # Create worksheet
    ws = workbook.add_worksheet(sheet_name)

    # Define formats
    fmt_header = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'bg_color': '#D9E1F2', 'border': 1})
    fmt_group = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'bg_color': '#D9E1F2', 'border': 1})
    fmt_grey = workbook.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True, 'bg_color': '#E7E6E6', 'border': 1})
    fmt_bold_center = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})
    fmt_center = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})
    fmt_left = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'text_wrap': True, 'border': 1})

    # Track column widths
    max_width_by_col = {}
    def update_width(col_idx: int, value: Any):
        if value is None:
            return
        try:
            text = str(value)
        except Exception:
            text = ""
        width = max(1, len(text))
        prev = max_width_by_col.get(col_idx, 0)
        if width > prev:
            max_width_by_col[col_idx] = width

    def sanitize(value: Any):
        """Convert NaN/Inf/None to safe Excel-friendly values."""
        # Pandas NA
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        # Numeric infinities
        try:
            if isinstance(value, (int,)):
                return value
            if isinstance(value, float):
                if not math.isfinite(value):
                    return ""
                # Prefer ints for integral floats
                if value.is_integer():
                    return int(value)
                return value
        except Exception:
            pass
        # Default passthrough
        return value

    # Header rows 1-3
    # Left columns (3 columns)
    for i, lbl in enumerate(left_columns):
        col = i
        for row in (0, 1, 2):
            v = lbl if row == 1 else None
            ws.write(row, col, v, fmt_header)
            update_width(col, v if v is not None else "")

    # After Source column index
    col_after_source = len(left_columns)

    # Event group/label/name columns (three header rows)
    ws.write(0, col_after_source, "Event Group:", fmt_header)
    ws.write(1, col_after_source, "Event Label:", fmt_header)
    ws.write(2, col_after_source, "Event Name:", fmt_header)
    update_width(col_after_source, "Event Group:")
    update_width(col_after_source, "Event Label:")
    update_width(col_after_source, "Event Name:")

    # RTSM column
    col_rtsm = col_after_source + 1
    for row in (0, 1, 2):
        ws.write(row, col_rtsm, "RTSM", fmt_header)
        update_width(col_rtsm, "RTSM")

    col_start_visits = col_rtsm + 1

    # Row 2: Event Label
    for j, vlabel in enumerate(visit_labels):
        c = col_start_visits + j
        if event_names[j] == "SCRN":
            event_label = "Screening"
        elif event_names[j] == "RAND":
            event_label = "Randomisation"
        elif "V" in event_names[j]:
            event_label = f"Visit {event_names[j][1:]}"
        elif "P" in event_names[j]:
            event_label = f"Phone Visit {event_names[j][1:]}"
        else:
            event_label = vlabel
        ws.write(1, c, event_label, fmt_bold_center)
        update_width(c, event_label)

    # Row 3: Event Name
    for j, ename in enumerate(event_names):
        c = col_start_visits + j
        ws.write(2, c, ename, fmt_bold_center)
        update_width(c, ename)

    # Row 1: Event Group merged ranges
    cur_group, group_start_col = None, None
    for j, g in enumerate(visit_groups):
        c = col_start_visits + j
        if cur_group is None:
            cur_group, group_start_col = g, c
        if g != cur_group:
            if c - 1 > group_start_col:
                ws.merge_range(0, group_start_col, 0, c - 1, cur_group, fmt_group)
            else:
                ws.write(0, group_start_col, cur_group, fmt_group)
            for cc in range(group_start_col, c):
                update_width(cc, cur_group)
            cur_group, group_start_col = g, c
    if group_start_col is not None:
        end_col = col_start_visits + num_visits - 1
        if end_col > group_start_col:
            ws.merge_range(0, group_start_col, 0, end_col, cur_group, fmt_group)
        else:
            ws.write(0, group_start_col, cur_group, fmt_group)
        for cc in range(group_start_col, col_start_visits + num_visits):
            update_width(cc, cur_group)

    # Extra headers after visits
    for idx, h in enumerate(extra_headers):
        c = col_start_visits + num_visits + idx
        ws.write(0, c, "", fmt_header)
        ws.write(1, c, h, fmt_bold_center)
        ws.write(2, c, "", fmt_header)
        update_width(c, h)

    # Blocks: Visit Dynamics + Event Window
    cur_row = 3
    dynamic_rows = [
        "Visit Dynamics (If Y, then Event should appear based on triggering criteria)",
        "Triggering: Event",
        "Triggering: Form",
        "Triggering: Item = Response (if specific response expected, else leave to accept any entered result)",
    ]
    event_window_rows = [
        "Assign Visit Window",
        "Offset Type (Previous Event, Specific Event, or None)",
        "Offset Days (Planned Visit Date, as calculated from Offset Event)",
        "Day Range - Early",
        "Day Range - Late",
    ]
    sections = [("Visit Dynamic Properties", dynamic_rows), ("Event Window Configuration", event_window_rows)]

    for section_title, attrs in sections:
        # Section header merged over left columns (avoid single-cell merge)
        end_col_left = len(left_columns) - 1
        if end_col_left > 0:
            ws.merge_range(cur_row, 0, cur_row, end_col_left, section_title, fmt_grey)
        else:
            ws.write(cur_row, 0, section_title, fmt_grey)
        for cc in range(0, len(left_columns)):
            update_width(cc, section_title)
        cur_row += 1

        for attr in attrs:
            # Label cell merged over left columns (avoid single-cell merge)
            end_col_left = len(left_columns) - 1
            if end_col_left > 0:
                ws.merge_range(cur_row, 0, cur_row, end_col_left, attr, fmt_grey)
            else:
                ws.write(cur_row, 0, attr, fmt_grey)
            update_width(0, attr)

            for j in range(num_visits):
                c = col_start_visits + j
                mapped_value = ""

                if attr.startswith("Visit Dynamics"):
                    eg = str(visit_groups[j]).lower()
                    if j >= rand_idx and ("end of treatment" not in eg and "end of study" not in eg):
                        mapped_value = "Y"
                elif attr.startswith("Triggering: Event"):
                    if event_names[j] == "RAND":
                        mapped_value = "SCRN"
                    elif event_names[j].startswith("V") and j > 0:
                        mapped_value = event_names[j - 1]
                    elif event_names[j].lower() == "follow-up":
                        mapped_value = "EOT"
                elif attr.startswith("Triggering: Form"):
                    mapped_value = ""
                    if event_names[j] == "RAND":
                        mapped_value = "ELIGIBILITY_CRITERIA"
                    elif j > rand_idx:
                        if "V" in visit_labels[j]:
                            if mapped_value == "":
                                mapped_value = "RANDOMISATION"
                elif attr.startswith("Assign Visit Window"):
                    mapped_value = "Y"
                elif attr.startswith("Offset Type") and "Offset Type" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Offset Type", "")
                elif attr.startswith("Offset Days") and "Offset Days" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Offset Days", "")
                elif attr.startswith("Day Range - Early") and "Day Range - Early" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Day Range - Early", "")
                elif attr.startswith("Day Range - Late") and "Day Range - Late" in df_visits.columns:
                    mapped_value = df_visits.iloc[j].get("Day Range - Late", "")

                if pd.isna(mapped_value):
                    mapped_value = ""
                v = sanitize(mapped_value)
                ws.write(cur_row, c, v, fmt_center)
                update_width(c, v)

            # RTSM column for this row
            ws.write(cur_row, col_rtsm, "", fmt_center)
            cur_row += 1

    # Forms table
    forms_start_row = cur_row

    # RTSM row
    ws.write(forms_start_row, 0, "RTSM", fmt_left)
    ws.write(forms_start_row, 1, "RTSM", fmt_left)
    ws.write(forms_start_row, 2, "Library", fmt_left)
    ws.write(forms_start_row, col_rtsm, "X", fmt_center)
    update_width(0, "RTSM"); update_width(1, "RTSM"); update_width(2, "Library"); update_width(col_rtsm, "X")
    for idx in range(len(extra_headers)):
        c = col_start_visits + num_visits + idx
        ws.write(forms_start_row, c, "", fmt_center)
    cur_row = forms_start_row + 1

    # Data rows from forms CSV, streamed in chunks
    chunksize = int(config.get('forms_csv_chunksize', 1000))
    for df_chunk in pd.read_csv(forms_csv, chunksize=chunksize):
        # normalize chunk column names
        df_chunk.columns = [str(c).strip() for c in df_chunk.columns]
        for _, r in df_chunk.iterrows():
            v0 = sanitize(r.get('Form Label', ''))
            v1 = sanitize(r.get('Form Name', ''))
            v2 = sanitize(r.get('Source', ''))
            ws.write(cur_row, 0, v0, fmt_left); update_width(0, v0)
            ws.write(cur_row, 1, v1, fmt_left); update_width(1, v1)
            ws.write(cur_row, 2, v2, fmt_left); update_width(2, v2)
            ws.write(cur_row, col_rtsm, "", fmt_center)

            for j, vlabel in enumerate(visit_labels):
                c = col_start_visits + j
                val = ""
                if vlabel in r.index:
                    val = r[vlabel]
                else:
                    en = event_names[j]
                    if en in r.index:
                        val = r[en]
                if pd.isna(val):
                    val = ""
                v = sanitize(val)
                ws.write(cur_row, c, v, fmt_center); update_width(c, v)

            # Extra columns
            extra_vals = {
                "Is Form Dynamic?": r.get("Is Form Dynamic?", "") or r.get("Is Form Dynamic", "") or r.get("IsDynamic", ""),
                "Form Dynamic Criteria": r.get("Form Dynamic Criteria", "") or r.get("Form Dynamic Criteria ", "")
            }
            for idx, colname in enumerate(extra_headers):
                c = col_start_visits + num_visits + idx
                val = sanitize(extra_vals.get(colname, ""))
                ws.write(cur_row, c, val, fmt_center); update_width(c, val)

            cur_row += 1

    # Freeze panes at forms_start_row and RTSM column
    ws.freeze_panes(forms_start_row, col_rtsm)

    # Set column widths (cap between 10 and 80, +3 padding)
    for col_idx, width in max_width_by_col.items():
        ws.set_column(col_idx, col_idx, max(10, min(80, width + 3)))

